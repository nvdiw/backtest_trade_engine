"""Parallel full-grid and budgeted adaptive optimization for ``ma_strategy``.

Examples:
    python optimize.py --auto -w 16
    python optimize.py --auto --resume -w 16
    python optimize.py --mode smart --tests 5000 -w 8
    python optimize.py --mode grid -w 8
"""

import argparse
import csv
import gzip
import itertools
import json
import math
import multiprocessing
import os
import random
import signal
import statistics
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

from ma_strategy import ma_strategy
from strategy_config import build_ma_strategy_config, load_ma_strategy_tune


# Every key is an existing MAStrategyConfig setting.  Defaults in
# strategy_config.py remain unchanged; the optimizer only passes candidates via
# ``tune`` and writes the winning values to JSON.
FULL_PARAM_GRID = {
    # Capital baseline (singletons prevent meaningless score scaling)
    "balance": [1000],
    "save_money": [0],
    "fee_rate": [0.0005],
    # Entry context and thresholds
    "entry_score_threshold": [6, 7, 8, 9, 10, 11, 12],
    "exit_score_threshold": [4, 5, 6, 7, 8, 9, 10],
    "ma_distance_threshold": [0.0008, 0.0010, 0.00125, 0.0015, 0.00159, 0.0020, 0.0025, 0.0030, 0.0040],
    "candle_move_threshold": [0.002, 0.004, 0.006, 0.008, 0.010, 0.012, 0.015],
    "impulse_move_threshold_pct": [0.75, 1.0, 1.5, 2.0, 2.5, 3.0],
    "impulse_lookback": [3, 4, 5, 6, 8, 10],
    "late_entry_atr_mult": [0.5, 0.8, 1.0, 1.2, 1.5, 2.0],
    "late_entry_body_ratio": [0.4, 0.6, 0.8, 1.0, 1.2],
    "late_entry_ema_pct": [0.002, 0.003, 0.004, 0.005, 0.007, 0.010],
    # Exit behavior
    "slope_window": [2, 3, 5, 8, 13],
    "trail_activate_pct": [0.003, 0.005, 0.007, 0.010, 0.015, 0.020],
    "trail_retrace_pct": [0.0015, 0.002, 0.003, 0.004, 0.005, 0.008],
    "loss_exit_pct_1": [0.015, 0.02, 0.025, 0.03, 0.04, 0.05],
    "loss_exit_pct_2": [0.03, 0.04, 0.05, 0.06, 0.08],
    "profit_exit_pct_1": [0.02, 0.03, 0.04, 0.05, 0.075, 0.10, 0.15],
    "profit_exit_pct_2": [0.06, 0.08, 0.10, 0.12, 0.15, 0.20],
    "loss_lock_step_pct": [0.005, 0.01, 0.015, 0.02, 0.03],
    "adx_exit_threshold": [12, 15, 18, 21, 24, 27],
    "adx_exit_lookback": [1, 2, 3, 5, 8],
    "opposite_atr_body_mult": [0.4, 0.6, 0.8, 1.0, 1.25, 1.5],
    "sharp_move_threshold_pct": [6, 8, 10, 12, 15, 18, 20],
    "sharp_move_lookback_candles": [100, 200, 300, 450, 600, 800, 1000],
    "post_cross_penalty_candles": [5, 10, 15, 20, 30],
    # Indicator periods and filters
    "ema_16_period": [10, 12, 14, 16, 18, 20, 24],
    "ma_50_period": [35, 40, 45, 50, 55, 60, 70],
    "ma_100_period": [80, 90, 100, 102, 110, 125, 140],
    "ma_200_period": [160, 180, 198, 200, 220, 240, 260],
    "period_adx": [8, 10, 12, 14, 16, 18, 21],
    "period_atr": [8, 10, 12, 14, 16, 18, 21],
    "period_atr_ma": [7, 10, 14, 18, 21, 28, 35],
    "period_vol_avg": [6, 8, 10, 12, 15, 18, 21, 30],
    "period_rsi": [7, 9, 11, 14, 18, 21],
    "entry_adx_threshold": [12, 15, 18, 20, 20.5, 22, 25, 28, 32],
    "entry_atr_threshold": [0.7, 0.85, 1.0, 1.1, 1.2, 1.35, 1.5],
    "volume_spike_multiplier": [0.9, 1.0, 1.1, 1.2, 1.24, 1.3, 1.45, 1.6, 1.8],
    "adx_filter": [False, True],
    "atr_filter": [False, True],
    "volume_filter": [False, True],
    # Score weights
    "entry_score_cross": [1, 2, 3, 4],
    "entry_score_ema_vs_ma50": [1, 2, 3, 4],
    "entry_score_ma_trend": [1, 2, 3],
    "entry_score_ma_distance_or_candle": [1, 2, 3],
    "entry_score_adx": [1, 2, 3],
    "entry_score_volume": [1, 2, 3],
    "entry_late_penalty": [0, 1, 2, 3, 4],
    "exit_score_loss_guard_1": [1, 2, 3, 4],
    "exit_score_loss_guard_2": [1, 2, 3, 4],
    "exit_score_profit_guard_1": [1, 2, 3, 4],
    "exit_score_profit_guard_2": [1, 2, 3, 4],
    "exit_score_ema_slope": [1, 2, 3, 4],
    "exit_score_ema_cross": [1, 2, 3, 4],
    "exit_score_ma_trend": [1, 2, 3],
    "exit_score_trailing": [1, 2, 3, 4],
    "exit_score_adx": [1, 2, 3],
    "exit_score_opposite_candle": [1, 2, 3],
    "post_cross_penalty_score": [0, 1, 2, 3, 4, 5],
    # Position sizing, safety, monthly controls, and scale-ins
    "trade_amount_percent": [0.2, 0.3, 0.4, 0.5, 0.6, 0.75, 0.9],
    "leverage": [1, 2, 3, 4, 5, 7, 10, 12],
    "safe_leverage_low": [1, 2, 3, 4],
    "safe_leverage_med": [2, 3, 4, 5, 6],
    "safe_leverage_high": [3, 4, 5, 6, 8],
    "safe_leverage_balance_pct_low": [60, 70, 75, 80],
    "safe_leverage_balance_pct_med": [70, 75, 80, 85],
    "safe_leverage_balance_pct_high": [80, 85, 90, 95],
    "save_money_recover_trigger_pct": [60, 65, 70, 75, 80],
    "max_open_trades": [1, 2, 3, 4, 5, 8],
    "cooldown_after_big_pnl": [0, 4, 8, 12, 24, 48, 96],
    "monthly_profit_percent_stop_trade": [5, 7, 9, 12, 15, 20],
    "monthly_loss_percent_stop_trade": [8, 12, 16, 19, 20, 25, 30],
    "monthly_compound": [0, 1, 2, 3, 5, 8],
    "monthly_profit_close_filter": [False, True],
    "monthly_loss_close_filter": [False, True],
    "consecutive_losses_month_stop_filter": [False, True],
    "consecutive_losses_stop_until_month": [3, 4, 5, 6, 8, 10],
    "skip_logic": [False, True],
    "scale_in_enabled": [False, True],
    "scale_entry_amount_percent": [0.05, 0.1, 0.15, 0.2, 0.3],
    "scale_entry_profit_trigger_pct": [0.005, 0.01, 0.02, 0.03, 0.039, 0.04, 0.06],
    "scale_entry_loss_trigger_pct": [0.005, 0.01, 0.02, 0.03, 0.04, 0.06],
    "scale_entry_on_profit_enabled": [False, True],
    "scale_entry_on_loss_enabled": [False, True],
    "profit_scale_entry_filter_enabled": [False, True],
    "profit_scale_entry_min_score": [2, 3, 4, 5, 6],
    "profit_scale_entry_atr_ratio_min": [0.8, 0.9, 1.0, 1.1, 1.25, 1.5],
    # RSI monthly sub-strategy used inside ma_strategy
    "rsi_trade_monthly_filter_on": [False, True],
    "rsi_long_open_monthly_profit": [10, 15, 20, 25, 30, 35],
    "rsi_long_close_monthly_profit": [60, 70, 79, 80, 85, 90],
    "rsi_short_open_monthly_profit": [55, 60, 65, 70, 75, 80],
    "rsi_short_close_monthly_profit": [10, 15, 20, 25, 30, 35],
    "rsi_long_tp_pct": [0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.10],
    "rsi_long_sl_pct": [0.02, 0.03, 0.04, 0.05, 0.06, 0.08],
    "rsi_short_tp_pct": [0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.10],
    "rsi_short_sl_pct": [0.01, 0.02, 0.03, 0.04, 0.05, 0.06],
    "rsi_max_open_trades": [1, 2, 3, 4],
    "rsi_trade_amount_percent": [0.1, 0.2, 0.3, 0.4, 0.5],
    "rsi_leverage": [1, 2, 3, 4, 5, 6, 8],
    "rsi_cooldown_filter": [False, True],
    "rsi_cooldown_bars": [0, 4, 8, 10, 12, 16, 24, 48],
    "lowest_rsi_last_n_value": [1, 2, 3, 5, 8, 13],
    "highest_rsi_last_n_value": [1, 2, 3, 5, 8, 13],
    "rsi_entry_buffer": [2, 4, 6, 8, 10, 12],
    "rsi_distance_threshold": [4, 6, 8, 10, 12, 15, 20],
    # Chart-only settings remain fixed because charts are disabled in optimize mode.
    "plot_max_candles": [1200],
    "plot_end_offset": [0],
    "plot_step_candles": [300],
    "plot_min_zoom_candles": [80],
    "plot_max_render_candles": [900],
    "plot_zoom_in_factor": [0.8],
    "plot_zoom_out_factor": [1.6],
    "plot_window_width_scale": [0.94],
    "plot_window_height_scale": [0.90],
    "plot_drag_preview_factor": [0.20],
    "plot_drag_update_interval_ms": [75],
    "plot_yscale_drag_sensitivity": [0.0030],
    "plot_post_cross_penalty_markers": [True],
}

FOCUSED_PARAM_GRID = {
    "entry_score_threshold": [6, 7, 8, 9, 10, 11, 12],
    "exit_score_threshold": [4, 5, 6, 7, 8, 9, 10],
    # Score weights
    "entry_score_cross": [1, 2, 3, 4],
    "entry_score_ema_vs_ma50": [1, 2, 3, 4],
    "entry_score_ma_trend": [1, 2, 3],
    "entry_score_ma_distance_or_candle": [1, 2, 3],
    "entry_score_adx": [1, 2, 3],
    "entry_score_volume": [1, 2, 3],
    "entry_late_penalty": [0, 1, 2, 3, 4],
    "exit_score_loss_guard_1": [1, 2, 3, 4],
    "exit_score_loss_guard_2": [1, 2, 3, 4],
    "exit_score_profit_guard_1": [1, 2, 3, 4],
    "exit_score_profit_guard_2": [1, 2, 3, 4],
    "exit_score_ema_slope": [1, 2, 3, 4],
    "exit_score_ema_cross": [1, 2, 3, 4],
    "exit_score_ma_trend": [1, 2, 3],
    "exit_score_trailing": [1, 2, 3, 4],
    "exit_score_adx": [1, 2, 3],
    "exit_score_opposite_candle": [1, 2, 3],
    "post_cross_penalty_score": [0, 1, 2, 3, 4, 5],
}


def _grid_subset(*prefixes, extra=()):
    keys = set(extra)
    for key in FULL_PARAM_GRID:
        if key.startswith(prefixes):
            keys.add(key)
    return {key: values for key, values in FULL_PARAM_GRID.items() if key in keys}


PARAMETER_PROFILES = {
    # The focused profile preserves the user's current score-weight search.
    "focused": FOCUSED_PARAM_GRID,
    "signal": _grid_subset(
        "entry_", "ma_distance", "candle_move", "impulse_", "late_entry_",
        "period_", "volume_spike", "adx_filter", "atr_filter", "volume_filter",
        extra=("ema_16_period", "ma_50_period", "ma_100_period", "ma_200_period"),
    ),
    "exit": _grid_subset(
        "exit_", "trail_", "loss_exit", "profit_exit", "loss_lock",
        "adx_exit", "opposite_", "sharp_move", "post_cross_", "slope_window",
    ),
    "risk": _grid_subset(
        "safe_leverage", "monthly_", "scale_", "profit_scale_", "consecutive_",
        extra=(
            "trade_amount_percent", "leverage", "save_money_recover_trigger_pct",
            "max_open_trades", "cooldown_after_big_pnl", "skip_logic",
        ),
    ),
    "rsi": _grid_subset("rsi_", "lowest_rsi_", "highest_rsi_"),
    "full": FULL_PARAM_GRID,
}

# Backward-compatible public name used by tests and imports.
param_grid = FULL_PARAM_GRID

RESULT_COLUMNS = [
    "final_balance_static", "final_balance_dynamic", "final_balance",
    "final_balance_without_fee",
    "total_profit", "realized_profit", "unrealized_profit", "open_positions",
    "total_fees", "saved_money", "liquidations", "total_profit_percent",
    "closed_trades", "wins", "losses", "long_trades", "long_wins",
    "long_losses", "short_trades", "short_wins", "short_losses",
    "maximum_drawdown", "win_rate", "profit_months", "loss_months",
    "score", "profit_factor", "expectancy_percent", "calmar_ratio",
    "rsi_total_trades", "rsi_wins", "rsi_losses", "rsi_winrate",
    "rsi_total_profit", "rsi_long_trades", "rsi_long_wins", "rsi_long_losses",
    "rsi_long_winrate", "rsi_long_profit", "rsi_short_trades", "rsi_short_wins",
    "rsi_short_losses", "rsi_short_winrate", "rsi_short_profit",
    "scale_total_trades", "scale_wins", "scale_losses", "scale_winrate",
    "scale_total_profit", "scale_long_trades", "scale_long_wins",
    "scale_long_losses", "scale_long_winrate", "scale_long_profit",
    "scale_short_trades", "scale_short_wins", "scale_short_losses",
    "scale_short_winrate", "scale_short_profit",
]

DERIVED_RESULT_COLUMNS = ["objective_score", "profit_per_trade"]
AUTO_TIME_COLUMNS = ["time_normalized_score", "range_candles"]
CANDLES_PER_YEAR_15M = 365.25 * 24 * 4
SURROGATE_MAX_TRAINING_SAMPLES = 1024
SURROGATE_CACHE_BOOTSTRAP_CYCLES = 24
SURROGATE_CACHE_VERSION = 1

_WORKER_START = "2025-01-01"
_WORKER_END = "2026-02-23"
_WORKER_BASE_TUNE = {}
DEFAULT_OUTPUT_DIR = os.path.join("outputs", "optimize")


def grid_size(grid):
    return math.prod(len(values) for values in grid.values())


def iter_grid_candidates(grid):
    """Yield the full Cartesian grid lazily without allocating it in memory."""
    keys = tuple(grid)
    for combo in itertools.product(*(grid[key] for key in keys)):
        yield dict(zip(keys, combo))


def _nearest_value(values, target):
    if target in values:
        return target
    if isinstance(target, (int, float)) and not isinstance(target, bool):
        numeric = [value for value in values if isinstance(value, (int, float))]
        if numeric:
            return min(numeric, key=lambda value: abs(value - target))
    return values[0]


def is_valid_candidate(candidate):
    """Reject combinations that violate basic parameter relationships."""
    ma_periods = [
        candidate.get("ema_16_period"), candidate.get("ma_50_period"),
        candidate.get("ma_100_period"), candidate.get("ma_200_period"),
    ]
    if all(value is not None for value in ma_periods) and ma_periods != sorted(ma_periods):
        return False
    leverage_tiers = [
        candidate.get("safe_leverage_low"), candidate.get("safe_leverage_med"),
        candidate.get("safe_leverage_high"), candidate.get("leverage"),
    ]
    if all(value is not None for value in leverage_tiers) and leverage_tiers != sorted(leverage_tiers):
        return False
    balance_tiers = [
        candidate.get("safe_leverage_balance_pct_low"),
        candidate.get("safe_leverage_balance_pct_med"),
        candidate.get("safe_leverage_balance_pct_high"),
    ]
    if all(value is not None for value in balance_tiers) and balance_tiers != sorted(balance_tiers):
        return False
    return True


class SmartCandidateGenerator:
    """Reproducible adaptive search over discrete or locally refined values."""

    def __init__(
        self,
        grid,
        seed=42,
        baseline_params=None,
        continuous_refinement=False,
        parameter_importance=None,
        refinement_round=1,
    ):
        if not grid or any(not values for values in grid.values()):
            raise ValueError("param_grid must contain at least one value per parameter")
        self.grid = {key: tuple(values) for key, values in grid.items()}
        self.keys = tuple(self.grid)
        self.mutable_keys = tuple(key for key, values in self.grid.items() if len(values) > 1)
        self.random = random.Random(seed)
        self.continuous_refinement = bool(continuous_refinement)
        self.refinement_round = max(1, int(refinement_round))
        supplied_importance = parameter_importance or {}
        self.parameter_importance = {
            key: max(0.01, float(supplied_importance.get(key, 1.0)))
            for key in self.mutable_keys
        }
        self.seen = set()
        self.local_queue = []
        self.local_queued = set()
        self.baseline_attempted = False
        defaults = build_ma_strategy_config(baseline_params)
        self.baseline = {
            key: _nearest_value(values, getattr(defaults, key, values[0]))
            for key, values in self.grid.items()
        }

    def _signature(self, candidate):
        return tuple(candidate[key] for key in self.keys)

    def _canonicalize(self, candidate):
        """Collapse inactive conditional parameters to avoid duplicate backtests."""
        candidate = dict(candidate)
        if candidate.get("scale_in_enabled") is False:
            for key in self.keys:
                if key.startswith(("scale_entry_", "profit_scale_entry_")):
                    candidate[key] = self.baseline[key]
        else:
            if candidate.get("scale_entry_on_profit_enabled") is False:
                for key in ("scale_entry_profit_trigger_pct",):
                    if key in candidate:
                        candidate[key] = self.baseline[key]
            if candidate.get("scale_entry_on_loss_enabled") is False:
                for key in ("scale_entry_loss_trigger_pct",):
                    if key in candidate:
                        candidate[key] = self.baseline[key]
            if candidate.get("profit_scale_entry_filter_enabled") is False:
                for key in (
                    "profit_scale_entry_min_score",
                    "profit_scale_entry_atr_ratio_min",
                ):
                    if key in candidate:
                        candidate[key] = self.baseline[key]
        if candidate.get("rsi_trade_monthly_filter_on") is False:
            for key in self.keys:
                if (
                    key.startswith(("rsi_", "lowest_rsi_", "highest_rsi_"))
                    and key != "rsi_trade_monthly_filter_on"
                ):
                    candidate[key] = self.baseline[key]
        elif candidate.get("rsi_cooldown_filter") is False:
            if "rsi_cooldown_bars" in candidate:
                candidate["rsi_cooldown_bars"] = self.baseline["rsi_cooldown_bars"]
        inactive_filter_settings = (
            ("adx_filter", ("entry_adx_threshold", "entry_score_adx")),
            ("atr_filter", ("entry_atr_threshold",)),
            ("volume_filter", ("volume_spike_multiplier", "entry_score_volume")),
            (
                "consecutive_losses_month_stop_filter",
                ("consecutive_losses_stop_until_month",),
            ),
        )
        for switch, dependent_keys in inactive_filter_settings:
            if candidate.get(switch) is False:
                for key in dependent_keys:
                    if key in candidate:
                        candidate[key] = self.baseline[key]
        return candidate

    def _random_candidate(self):
        return {key: self.random.choice(values) for key, values in self.grid.items()}

    @staticmethod
    def _is_numeric_values(values):
        return all(
            isinstance(value, (int, float)) and not isinstance(value, bool)
            for value in values
        )

    def _numeric_refinement_step(self, values):
        ordered = sorted(set(values))
        if all(isinstance(value, int) and not isinstance(value, bool) for value in ordered):
            return 1
        gaps = [
            right - left for left, right in zip(ordered, ordered[1:])
            if right > left
        ]
        if not gaps:
            return 0
        # Each completed auto cycle can halve the smallest coarse-grid gap. The
        # cap avoids creating meaningless floating-point precision indefinitely.
        divisor = 2 ** min(4, self.refinement_round)
        return min(gaps) / divisor

    @staticmethod
    def _float_precision(values, step):
        def decimal_places(value):
            text = f"{float(value):.12f}".rstrip("0")
            return len(text.partition(".")[2])

        return min(12, max([decimal_places(value) for value in values] + [decimal_places(step)]))

    def _refined_neighbors(self, key, current):
        values = self.grid[key]
        if not self.continuous_refinement or not self._is_numeric_values(values):
            ordered = list(values)
            if current in ordered:
                index = ordered.index(current)
            else:
                index = min(range(len(ordered)), key=lambda i: abs(ordered[i] - current))
            return [
                ordered[neighbor_index]
                for neighbor_index in (index - 1, index + 1)
                if 0 <= neighbor_index < len(ordered)
            ]

        lower, upper = min(values), max(values)
        step = self._numeric_refinement_step(values)
        if step <= 0:
            return []
        precision = self._float_precision(values, step)
        neighbors = []
        for direction in (-1, 1):
            value = current + direction * step
            value = max(lower, min(upper, value))
            if all(isinstance(item, int) and not isinstance(item, bool) for item in values):
                value = int(round(value))
            else:
                value = round(value, precision)
            if value != current and value not in neighbors:
                neighbors.append(value)
        return neighbors

    def _weighted_mutation_keys(self, count):
        available = list(self.mutable_keys)
        chosen = []
        while available and len(chosen) < count:
            weights = [self.parameter_importance.get(key, 1.0) for key in available]
            key = self.random.choices(available, weights=weights, k=1)[0]
            available.remove(key)
            chosen.append(key)
        return chosen

    def _mutate_key(self, candidate, key, prefer_local=True):
        values = self.grid[key]
        neighbors = self._refined_neighbors(key, candidate[key])
        if prefer_local and neighbors and self.random.random() < 0.85:
            candidate[key] = self.random.choice(neighbors)
        else:
            candidate[key] = self.random.choice(values)

    def _elite_choice(self, elites):
        # Rank weighting prevents one early lucky candidate from monopolizing search.
        weights = list(range(len(elites), 0, -1))
        return self.random.choices(elites, weights=weights, k=1)[0]

    def _guided_candidate(self, elites, progress=0.0, crossover_probability=0.20):
        # Occasionally cross two good candidates, then mutate. Mutation becomes
        # narrower as the budget is consumed (exploration -> exploitation).
        parent = self._elite_choice(elites)
        candidate = {key: parent["params"][key] for key in self.keys}
        if len(elites) > 1 and self.random.random() < crossover_probability:
            other = self._elite_choice(elites)["params"]
            for key in self.mutable_keys:
                if self.random.random() < 0.5:
                    candidate[key] = other[key]

        max_mutations = max(2, round(math.sqrt(max(1, len(self.mutable_keys)))))
        mutation_count = max(1, round(max_mutations * (1.0 - 0.70 * progress)))
        mutation_count = min(len(self.mutable_keys), mutation_count)
        for key in self._weighted_mutation_keys(mutation_count):
            self._mutate_key(candidate, key, prefer_local=True)
        return candidate

    def _crossover_candidate(self, elites, progress=0.0):
        candidate = self._guided_candidate(
            elites, progress=progress, crossover_probability=1.0
        )
        return candidate

    def _queue_elite_neighbors(self, elites):
        """Queue deterministic one-step neighbors around the current elites."""
        for elite in elites:
            parent = {key: elite["params"][key] for key in self.keys}
            ordered_keys = sorted(
                self.mutable_keys,
                key=lambda key: self.parameter_importance.get(key, 1.0),
                reverse=True,
            )
            for key in ordered_keys:
                current = parent[key]
                for neighbor in self._refined_neighbors(key, current):
                    candidate = dict(parent)
                    candidate[key] = neighbor
                    candidate = self._canonicalize(candidate)
                    signature = self._signature(candidate)
                    if (
                        signature not in self.seen
                        and signature not in self.local_queued
                        and is_valid_candidate(candidate)
                    ):
                        self.local_queue.append(candidate)
                        self.local_queued.add(signature)

    def generate(self, count, elites=None, progress=0.0, progress_label=None):
        candidates = []
        attempts = 0
        max_attempts = max(1000, count * 100)
        if elites:
            self._queue_elite_neighbors(elites)
        while len(candidates) < count and attempts < max_attempts:
            attempts += 1
            if not self.baseline_attempted:
                candidate = dict(self.baseline)
                self.baseline_attempted = True
            elif self.local_queue and self.random.random() < (0.35 + 0.50 * progress):
                candidate = self.local_queue.pop(0)
                self.local_queued.discard(self._signature(candidate))
            elif elites and self.random.random() >= max(0.12, 0.40 * (1.0 - progress)):
                candidate = self._guided_candidate(elites, progress=progress)
            else:
                candidate = self._random_candidate()
            candidate = self._canonicalize(candidate)
            signature = self._signature(candidate)
            if signature in self.seen or not is_valid_candidate(candidate):
                continue
            self.seen.add(signature)
            candidates.append(candidate)
            if progress_label and (
                len(candidates) == count
                or len(candidates) % max(1, math.ceil(count / 20)) == 0
            ):
                _show_loading_progress(progress_label, len(candidates), count)
        return candidates

    def generate_auto(self, count, elites=None, progress_label=None):
        """Generate the 25% exploration / 15% crossover / 60% local mix."""
        if not elites:
            return self.generate(count, progress_label=progress_label)
        self._queue_elite_neighbors(elites)
        candidates = []
        attempts = 0
        max_attempts = max(2000, count * 200)
        while len(candidates) < count and attempts < max_attempts:
            attempts += 1
            roll = self.random.random()
            if roll < 0.25:
                candidate = self._random_candidate()
            elif roll < 0.40:
                candidate = self._crossover_candidate(elites, progress=0.85)
            elif self.local_queue and self.random.random() < 0.50:
                candidate = self.local_queue.pop(0)
                self.local_queued.discard(self._signature(candidate))
            else:
                candidate = self._guided_candidate(
                    elites, progress=0.85, crossover_probability=0.0
                )
            candidate = self._canonicalize(candidate)
            signature = self._signature(candidate)
            if signature in self.seen or not is_valid_candidate(candidate):
                continue
            self.seen.add(signature)
            candidates.append(candidate)
            if progress_label and (
                len(candidates) == count
                or len(candidates) % max(1, math.ceil(count / 20)) == 0
            ):
                _show_loading_progress(progress_label, len(candidates), count)
        return candidates


def _parse_bound(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def _init_worker(start, end, base_tune=None, ignore_keyboard_interrupt=False):
    global _WORKER_START, _WORKER_END, _WORKER_BASE_TUNE
    if ignore_keyboard_interrupt:
        signal.signal(signal.SIGINT, signal.SIG_IGN)
    _WORKER_START = start
    _WORKER_END = end
    _WORKER_BASE_TUNE = dict(base_tune or {})
    # Warm the largest repeated I/O cost once per process.
    from trade_engine import TradeEngine
    TradeEngine.load_market_data(start=start, end=end)


def _evaluate_candidate(index, params, base_tune, start, end):
    started = time.perf_counter()
    try:
        result = ma_strategy(
            tune={**base_tune, **params, "optimize": True},
            start=start,
            end=end,
        )
        error = None
    except Exception as exc:  # return errors to the parent without killing the run
        result = None
        error = f"{type(exc).__name__}: {exc}"
    return index, params, result, time.perf_counter() - started, error


def _evaluate_task(task):
    """Small multiprocessing payload: workers merge the shared base tune locally."""
    index, params = task
    return _evaluate_candidate(
        index, params, _WORKER_BASE_TUNE, _WORKER_START, _WORKER_END
    )


def _score(result):
    if not result:
        return -math.inf
    value = result.get("score", -math.inf)
    try:
        value = float(value)
    except (TypeError, ValueError):
        return -math.inf
    return value if math.isfinite(value) else -math.inf


def _objective_score(result, min_trades=0, max_drawdown=None):
    """Apply research-quality constraints before a result can become elite."""
    value = _score(result)
    if not result or value == -math.inf:
        return -math.inf
    if int(result.get("closed_trades", 0) or 0) < min_trades:
        return -math.inf
    if max_drawdown is not None:
        raw_drawdown = result.get("maximum_drawdown")
        if raw_drawdown is None or abs(float(raw_drawdown)) > max_drawdown:
            return -math.inf
    return value


def _robust_validation_score(
    train_result,
    validation_result,
    overfit_penalty=0.25,
    train_candles=None,
    validation_candles=None,
):
    """Prefer strong validation performance and penalize train-only excess."""
    train_score = _score(train_result)
    validation_score = _score(validation_result)
    if train_candles is not None and validation_candles is not None:
        normalized_train = _time_normalized_score(train_score, train_candles)
        normalized_validation = _time_normalized_score(
            validation_score, validation_candles
        )
        if normalized_train is not None:
            train_score = normalized_train
        if normalized_validation is not None:
            validation_score = normalized_validation
    if not math.isfinite(validation_score):
        return -math.inf
    optimistic_gap = max(0.0, train_score - validation_score)
    return validation_score - overfit_penalty * optimistic_gap


def _time_normalized_score(score, range_candles):
    """Convert a raw optimizer score to a comparable 15-minute annual rate."""
    numeric_score = _finite_number(score)
    try:
        candles = int(range_candles)
    except (TypeError, ValueError):
        return None
    if numeric_score is None or candles <= 0:
        return None
    return numeric_score * CANDLES_PER_YEAR_15M / candles


def _range_candle_count(range_start, range_end):
    return max(1, _bound_index(range_end) - _bound_index(range_start))


AUTO_STAGE_ORDER = ("discovery", "validation", "stress", "walk_forward", "final")
AUTO_STAGE_WEIGHTS = {
    "discovery": 0.25,
    "validation": 0.20,
    "stress": 0.15,
    "walk_forward": 0.30,
    "final": 0.10,
}
AUTO_STATE_VERSION = 2
LEGACY_AUTO_STATE_VERSIONS = {1, AUTO_STATE_VERSION}


def _finite_number(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _record_comparable_score(record):
    normalized = _finite_number(record.get("time_normalized_score"))
    return normalized if normalized is not None else _finite_number(
        record.get("objective_score")
    )


def _importance_target(record, target):
    if target == "objective_score":
        return _finite_number(record.get("objective_score"))
    return _finite_number((record.get("result") or {}).get(target))


def _learn_parameter_importance(records, parameter_keys, target="objective_score"):
    """Estimate which parameters explain the largest share of result variance.

    Exact values are grouped for small discrete spaces. Highly varied numeric
    parameters are binned so locally refined off-grid values remain useful.
    A small exploration floor prevents an early noisy estimate from permanently
    freezing any parameter.
    """
    usable = []
    for record in records:
        value = _importance_target(record, target)
        if value is not None:
            usable.append((record["params"], value))
    if len(usable) < 4:
        equal = 1.0 / max(1, len(parameter_keys))
        return {
            key: {"weight": equal, "effect": 0.0, "groups": 0, "samples": len(usable)}
            for key in parameter_keys
        }

    targets = [value for _, value in usable]
    overall_mean = statistics.fmean(targets)
    total_variance = statistics.fmean((value - overall_mean) ** 2 for value in targets)
    raw = {}
    for key in parameter_keys:
        key_values = [params[key] for params, _ in usable]
        unique = list(dict.fromkeys(key_values))
        numeric = all(
            isinstance(value, (int, float)) and not isinstance(value, bool)
            for value in unique
        )
        groups = {}
        if numeric and len(unique) > 12:
            lower, upper = min(unique), max(unique)
            width = (upper - lower) / 8 if upper > lower else 0
            for (params, target_value) in usable:
                group = 0 if width == 0 else min(7, int((params[key] - lower) / width))
                groups.setdefault(group, []).append(target_value)
        else:
            for params, target_value in usable:
                groups.setdefault((type(params[key]).__name__, params[key]), []).append(target_value)

        between = sum(
            len(values) * (statistics.fmean(values) - overall_mean) ** 2
            for values in groups.values()
        ) / len(usable)
        effect = between / total_variance if total_variance > 0 else 0.0
        confidence = min(1.0, len(usable) / max(20.0, len(groups) * 4.0))
        raw[key] = {
            "effect": max(0.0, min(1.0, effect)) * confidence,
            "groups": len(groups),
            "samples": len(usable),
        }

    # The floor reserves exploration for every variable while high-effect
    # variables receive most local mutations and deterministic neighbor tests.
    scores = {key: 0.05 + item["effect"] for key, item in raw.items()}
    total = sum(scores.values()) or 1.0
    return {
        key: {**raw[key], "weight": scores[key] / total}
        for key in parameter_keys
    }


def _smooth_parameter_importance(previous, learned, previous_weight=0.65):
    if not previous:
        return learned
    blended = {}
    for key, item in learned.items():
        old = previous.get(key, {})
        weight = previous_weight * float(old.get("weight", 0.0)) + (
            1.0 - previous_weight
        ) * float(item.get("weight", 0.0))
        blended[key] = {**item, "weight": weight}
    total = sum(item["weight"] for item in blended.values()) or 1.0
    for item in blended.values():
        item["weight"] /= total
    return blended


class ExtraTreesSurrogate:
    """Small dependency-free extremely-randomized tree ensemble.

    It is intentionally limited to the numeric/boolean parameter spaces used by
    this optimizer.  The ensemble predicts both a mean score and disagreement
    between trees, which lets auto mode balance exploitation and exploration.
    """

    def __init__(
        self, n_trees=32, max_depth=10, min_leaf=4, max_features=None, seed=42,
    ):
        self.n_trees = max(1, int(n_trees))
        self.max_depth = max(1, int(max_depth))
        self.min_leaf = max(1, int(min_leaf))
        self.max_features = max_features
        self.seed = int(seed)
        self.trees = []

    @staticmethod
    def _leaf(targets, indices):
        return ("leaf", statistics.fmean(targets[index] for index in indices))

    def _build_tree(self, features, targets, indices, depth, randomizer):
        if depth >= self.max_depth or len(indices) < self.min_leaf * 2:
            return self._leaf(targets, indices)
        node_targets = [targets[index] for index in indices]
        if max(node_targets) - min(node_targets) <= 1e-12:
            return self._leaf(targets, indices)

        feature_count = len(features[0])
        requested = self.max_features or max(1, round(math.sqrt(feature_count)))
        selected_features = randomizer.sample(
            range(feature_count), min(feature_count, requested)
        )
        best = None
        for feature_index in selected_features:
            values = [features[index][feature_index] for index in indices]
            lower, upper = min(values), max(values)
            if lower == upper:
                continue
            for _ in range(3):
                threshold = randomizer.uniform(lower, upper)
                left = []
                right = []
                for index in indices:
                    target = (
                        left
                        if features[index][feature_index] <= threshold
                        else right
                    )
                    target.append(index)
                if len(left) < self.min_leaf or len(indices) - len(left) < self.min_leaf:
                    continue
                left_mean = statistics.fmean(targets[index] for index in left)
                right_mean = statistics.fmean(targets[index] for index in right)
                loss = sum((targets[index] - left_mean) ** 2 for index in left)
                loss += sum((targets[index] - right_mean) ** 2 for index in right)
                if best is None or loss < best[0]:
                    best = (loss, feature_index, threshold, left, right)
        if best is None:
            return self._leaf(targets, indices)
        _, feature_index, threshold, left, right = best
        return (
            "node", feature_index, threshold,
            self._build_tree(features, targets, left, depth + 1, randomizer),
            self._build_tree(features, targets, right, depth + 1, randomizer),
        )

    def fit(self, features, targets, progress_label=None):
        if not features or len(features) != len(targets):
            raise ValueError("surrogate training features and targets must be non-empty")
        sample_count = len(features)
        self.trees = []
        for tree_index in range(self.n_trees):
            randomizer = random.Random(self.seed + tree_index * 104729)
            # Random subsampling gives useful model disagreement without letting
            # duplicate bootstrap rows dominate small optimization histories.
            subset_size = max(self.min_leaf * 2, round(sample_count * 0.80))
            subset_size = min(sample_count, subset_size)
            indices = randomizer.sample(range(sample_count), subset_size)
            self.trees.append(
                self._build_tree(features, targets, indices, 0, randomizer)
            )
            if progress_label and (
                tree_index + 1 == self.n_trees
                or (tree_index + 1) % max(1, math.ceil(self.n_trees / 20)) == 0
            ):
                _show_loading_progress(progress_label, tree_index + 1, self.n_trees)
        return self

    @staticmethod
    def _predict_tree(tree, features):
        while tree[0] == "node":
            _, feature_index, threshold, left, right = tree
            tree = left if features[feature_index] <= threshold else right
        return tree[1]

    def predict_mean_std(self, feature_rows, progress_label=None):
        if not self.trees:
            raise ValueError("surrogate must be fitted before prediction")
        output = []
        total = len(feature_rows)
        for index, row in enumerate(feature_rows, start=1):
            predictions = [self._predict_tree(tree, row) for tree in self.trees]
            output.append((
                statistics.fmean(predictions),
                statistics.pstdev(predictions) if len(predictions) > 1 else 0.0,
            ))
            if progress_label and (
                index == total or index % max(1, math.ceil(total / 20)) == 0
            ):
                _show_loading_progress(progress_label, index, total)
        return output


def _record_percentiles(records):
    valid = [
        record for record in records
        if _record_comparable_score(record) is not None
    ]
    valid.sort(
        key=lambda record: _record_comparable_score(record),
        reverse=True,
    )
    denominator = max(1, len(valid) - 1)
    return {
        record["candidate_id"]: 1.0 - rank / denominator
        for rank, record in enumerate(valid)
    }


def _combine_auto_stage_records(stage_records):
    """Rank candidates across every completed, non-overlapping market regime."""
    completed_stages = [stage for stage in AUTO_STAGE_ORDER if stage in stage_records]
    if not completed_stages:
        return []
    final_stage = completed_stages[-1]
    record_maps = {
        stage: {record["candidate_id"]: record for record in records}
        for stage, records in stage_records.items()
    }
    percentiles = {
        stage: _record_percentiles(records) for stage, records in stage_records.items()
    }
    stage_weights = AUTO_STAGE_WEIGHTS
    if "walk_forward" not in completed_stages:
        # Preserve the ranking contract used by version-1 campaigns while an
        # interrupted legacy cycle is being finished.
        stage_weights = {
            "discovery": 0.40, "validation": 0.30,
            "stress": 0.20, "final": 0.10,
        }
    combined = []
    for latest in stage_records[final_stage]:
        candidate_id = latest["candidate_id"]
        if any(candidate_id not in record_maps[stage] for stage in completed_stages):
            continue
        ranks = []
        weighted_total = 0.0
        weight_total = 0.0
        stage_scores = {}
        stage_metrics = {}
        qualified = True
        for stage in completed_stages:
            record = record_maps[stage][candidate_id]
            score = _finite_number(record.get("objective_score"))
            normalized_score = _record_comparable_score(record)
            percentile = percentiles[stage].get(candidate_id)
            if score is None or normalized_score is None or percentile is None:
                qualified = False
                break
            weight = stage_weights[stage]
            ranks.append(percentile)
            weighted_total += weight * percentile
            weight_total += weight
            stage_scores[stage] = score
            stage_metrics[stage] = {
                **record["result"],
                "time_normalized_score": normalized_score,
            }
        if not qualified:
            robust_score = -math.inf
        else:
            weighted_rank = weighted_total / weight_total
            dispersion = statistics.pstdev(ranks) if len(ranks) > 1 else 0.0
            transformed_quality = sum(
                stage_weights[stage]
                * math.copysign(
                    math.log1p(abs(stage_metrics[stage]["time_normalized_score"])),
                    stage_metrics[stage]["time_normalized_score"],
                )
                for stage in completed_stages
            ) / weight_total
            robust_score = (
                50.0 * weighted_rank
                + 20.0 * min(ranks)
                - 10.0 * dispersion
                + 10.0 * transformed_quality
            )
        combined.append({
            "candidate_id": candidate_id,
            "params": latest["params"],
            "robust_score": robust_score,
            "worst_stage_percentile": min(ranks) if ranks else None,
            "stage_scores": stage_scores,
            "stage_metrics": stage_metrics,
        })
    combined.sort(key=lambda record: record["robust_score"], reverse=True)
    return combined


def _latest_market_end():
    """Return an exclusive timestamp immediately after the last valid candle."""
    from get_candle_index import _open_times

    open_times = _open_times().dropna()
    if open_times.empty:
        raise ValueError("market data does not contain a valid Open time")
    recent = open_times.tail(100).sort_values()
    deltas = recent.diff().dropna()
    candle_delta = deltas.median() if not deltas.empty else timedelta(minutes=15)
    if candle_delta <= timedelta(0):
        candle_delta = timedelta(minutes=15)
    return (open_times.iloc[-1] + candle_delta).strftime("%Y-%m-%d %H:%M:%S")


def _timestamp_now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _show_loading_progress(label, completed, total):
    """Render one in-place startup progress line for potentially large histories."""
    if total < 5:
        return
    percent = 100.0 * completed / max(1, total)
    end = "\n" if completed >= total else ""
    print(
        f"\r{label}: {percent:6.2f}% ({completed:,}/{total:,})",
        end=end,
        flush=True,
    )


def _json_safe(value):
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if hasattr(value, "item"):
        return _json_safe(value.item())
    return value


def _replace_with_retry(temporary, path, attempts=20):
    """Replace a file atomically, tolerating short-lived Windows file locks."""
    for attempt in range(attempts):
        try:
            os.replace(temporary, path)
            return
        except PermissionError:
            if attempt + 1 >= attempts:
                raise
            # Editors, indexers, antivirus, and sync clients can briefly open a
            # destination without FILE_SHARE_DELETE, which makes os.replace fail.
            time.sleep(min(0.05 * (2 ** attempt), 1.0))


def _write_json(path, payload):
    path = Path(path)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(_json_safe(payload), file, indent=2, ensure_ascii=False)
        file.write("\n")
    _replace_with_retry(temporary, path)


def _save_optimizer_workbook(results_path, output_path, parameter_keys, max_rows=5000):
    """Create a filterable, frozen-header XLSX companion to the raw CSV."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return None

    results_path = Path(results_path)
    if not results_path.is_file():
        return None
    max_rows = max(1, int(max_rows))
    ranked = None
    for chunk in pd.read_csv(results_path, chunksize=50_000):
        if chunk.empty:
            continue
        score_column = "objective_score" if "objective_score" in chunk else "score"
        chunk = chunk.sort_values(score_column, ascending=False, na_position="last").head(max_rows)
        ranked = chunk if ranked is None else pd.concat([ranked, chunk], ignore_index=True)
        ranked = ranked.sort_values(
            score_column, ascending=False, na_position="last"
        ).head(max_rows)
    if ranked is None or ranked.empty:
        return None

    def existing(columns):
        return [column for column in columns if column in ranked.columns]

    identity = existing(["test_index", "score", "objective_score", "duration_s"])
    sheets = {
        "Rankings": ranked,
        "Parameters": ranked[identity + existing(parameter_keys)],
        "RSI Metrics": ranked[identity + [
            column for column in ranked.columns if column.startswith("rsi_")
        ]],
        "Scale Metrics": ranked[identity + [
            column for column in ranked.columns if column.startswith("scale_")
        ]],
    }
    core_columns = existing([
        "test_index", "score", "objective_score", "final_balance",
        "final_balance_without_fee", "total_profit", "realized_profit",
        "unrealized_profit", "open_positions", "total_fees", "saved_money",
        "liquidations", "long_trades", "short_trades",
        "total_profit_percent", "closed_trades", "wins", "losses", "win_rate",
        "maximum_drawdown", "profit_factor", "expectancy_percent", "calmar_ratio",
        "profit_per_trade", "duration_s",
    ])
    sheets["Core Metrics"] = ranked[core_columns]

    output_path = Path(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_frame in sheets.items():
            sheet_frame.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for table_index, worksheet in enumerate(workbook.worksheets, start=1):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 24
        for column_index, cells in enumerate(worksheet.columns, start=1):
            values = [str(cell.value or "") for cell in cells[:200]]
            width = min(42, max(10, max(map(len, values), default=10) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            table = Table(
                displayName=f"OptimizerTable{table_index}", ref=worksheet.dimensions
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
                showLastColumn=False, showColumnStripes=False,
            )
            worksheet.add_table(table)
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for metric in ("score", "objective_score", "total_profit", "win_rate"):
            column = headers.get(metric)
            if column and worksheet.max_row >= 3:
                letter = get_column_letter(column)
                worksheet.conditional_formatting.add(
                    f"{letter}2:{letter}{worksheet.max_row}",
                    ColorScaleRule(
                        start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B",
                    ),
                )
    workbook.save(output_path)
    return output_path


def _write_best_files(output_dir, best, mode, requested_tests, completed, elapsed, seed,
                      metadata=None):
    if best is None:
        return
    _write_json(Path(output_dir) / "best_params.json", best["params"])
    summary = {
        "mode": mode,
        "requested_tests": requested_tests,
        "completed_tests": completed,
        "seed": seed,
        "elapsed_seconds": round(elapsed, 3),
        "best_test_index": best["index"],
        "best_duration_seconds": round(best["duration"], 4),
        "best_params": best["params"],
        "best_metrics": best["result"],
    }
    if metadata:
        summary.update(metadata)
    _write_json(Path(output_dir) / "optimization_summary.json", summary)


def _result_row(keys, index, params, result, duration, objective_score=None):
    row = {"test_index": index, "duration_s": round(duration, 4)}
    row.update({key: params[key] for key in keys})
    row.update({key: result.get(key) for key in RESULT_COLUMNS})
    closed_trades = int(result.get("closed_trades", 0) or 0)
    realized_profit = result.get("realized_profit")
    if realized_profit is None:
        realized_profit = result.get("total_profit", 0)
    realized_profit = float(realized_profit or 0)
    row["objective_score"] = objective_score
    row["profit_per_trade"] = (
        realized_profit / closed_trades if closed_trades else None
    )
    return row


def _result_fieldnames(keys):
    """Put outcome and trade metrics before the usually much wider parameter set."""
    metrics = [column for column in RESULT_COLUMNS if column != "score"]
    return [
        "test_index", "objective_score", "score", *metrics,
        "profit_per_trade", "duration_s", *keys,
    ]


def _parse_grid_csv_value(raw, values):
    for value in values:
        if str(value) == raw:
            return value
    raise ValueError(f"saved value {raw!r} is not present in the current grid")


def _parse_metric(raw):
    if raw in (None, ""):
        return None
    try:
        number = float(raw)
    except (TypeError, ValueError):
        return raw
    return int(number) if number.is_integer() else number


def _read_resume_records(path, grid, base_tune):
    records = []
    with Path(path).open(newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        # New metric columns are optional so older compatible checkpoints can
        # still resume after the reporting schema grows.
        required = {"test_index", "duration_s", "score", *grid}
        missing = required - set(reader.fieldnames or ())
        if missing:
            raise ValueError(
                "cannot resume because CSV columns do not match this profile: "
                + ", ".join(sorted(missing))
            )
        for row in reader:
            selected = {
                key: _parse_grid_csv_value(row[key], grid[key]) for key in grid
            }
            result = {key: _parse_metric(row.get(key)) for key in RESULT_COLUMNS}
            records.append({
                "index": int(row["test_index"]),
                "params": {**base_tune, **selected},
                "result": result,
                "duration": float(row["duration_s"]),
            })
    return records


def _load_base_tune(args):
    source = getattr(args, "base_source", "config")
    if source == "config":
        return {}, "strategy_config.py"
    path = Path(getattr(args, "base_params", "outputs/optimize/best_params.json"))
    if not path.is_file():
        raise FileNotFoundError(
            f"base parameter file not found: {path}. Run an optimization first or "
            "use --base-source=config."
        )
    return load_ma_strategy_tune(path), str(path)


def _validate_top_candidates(records, args, workers, chunksize):
    validation_start = getattr(args, "validation_start", None)
    validation_end = getattr(args, "validation_end", None)
    if not validation_start and not validation_end:
        return None, []
    if not validation_start or not validation_end:
        raise ValueError("--validation-start and --validation-end must be used together")

    top_count = min(getattr(args, "validation_top", 20), len(records))
    candidates = records[:top_count]
    start = _parse_bound(validation_start)
    end = _parse_bound(validation_end)
    train_candles = _range_candle_count(args.start, args.end)
    validation_candles = _range_candle_count(validation_start, validation_end)
    tasks = [(record["index"], record["params"]) for record in candidates]
    if workers > 1:
        pool = multiprocessing.Pool(
            workers, initializer=_init_worker, initargs=(start, end, {}, True)
        )
        try:
            evaluated = list(pool.imap_unordered(_evaluate_task, tasks, chunksize=chunksize))
        except KeyboardInterrupt:
            pool.terminate()
            pool.join()
            raise
        else:
            pool.close()
            pool.join()
    else:
        evaluated = [
            _evaluate_candidate(index, params, {}, start, end)
            for index, params in tasks
        ]

    train_by_index = {record["index"]: record for record in candidates}
    validated = []
    for index, params, result, duration, error in evaluated:
        train = train_by_index[index]
        validation_qualified = (
            not error and math.isfinite(_objective_score(
                result,
                min_trades=getattr(args, "min_trades", 0),
                max_drawdown=getattr(args, "max_drawdown", None),
            ))
        )
        robust_score = (
            -math.inf if not validation_qualified else _robust_validation_score(
                train["result"], result, getattr(args, "overfit_penalty", 0.25),
                train_candles=train_candles,
                validation_candles=validation_candles,
            )
        )
        validated.append({
            "index": index,
            "params": params,
            "result": result,
            "duration": duration,
            "error": error,
            "training_result": train["result"],
            "training_time_normalized_score": _time_normalized_score(
                _score(train["result"]), train_candles
            ),
            "validation_time_normalized_score": _time_normalized_score(
                _score(result), validation_candles
            ),
            "robust_score": robust_score,
        })
    validated.sort(key=lambda item: item["robust_score"], reverse=True)
    return (validated[0] if validated else None), validated


def _auto_ranges(args, resolved_end):
    return {
        "discovery": [args.auto_discovery_start, resolved_end],
        "validation": [args.auto_validation_start, args.auto_discovery_start],
        "stress": [args.auto_stress_start, args.auto_validation_start],
        "final": [args.auto_stress_start, resolved_end],
    }


def _bound_index(value):
    parsed = _parse_bound(value)
    if isinstance(parsed, int):
        return parsed
    from get_candle_index import get_candle_index

    return int(get_candle_index(parsed))


def _format_range_bound(value):
    """Return a human-readable market timestamp for a date or candle bound."""
    parsed = _parse_bound(value)
    if not isinstance(parsed, int):
        try:
            return parsed.strftime("%Y-%m-%d %H:%M")
        except (AttributeError, ValueError):
            return str(value)

    try:
        from get_candle_index import _open_times

        open_times = _open_times().dropna()
        if open_times.empty:
            return str(value)
        if 0 <= parsed < len(open_times):
            timestamp = open_times.iloc[parsed]
        elif parsed == len(open_times):
            recent = open_times.tail(100).sort_values()
            deltas = recent.diff().dropna()
            candle_delta = deltas.median() if not deltas.empty else timedelta(minutes=15)
            if candle_delta <= timedelta(0):
                candle_delta = timedelta(minutes=15)
            timestamp = open_times.iloc[-1] + candle_delta
        else:
            return str(value)
        return timestamp.strftime("%Y-%m-%d %H:%M")
    except (IndexError, OSError, TypeError, ValueError):
        return str(value)


def _validate_auto_ranges(ranges):
    discovery_start = _bound_index(ranges["discovery"][0])
    validation_start = _bound_index(ranges["validation"][0])
    stress_start = _bound_index(ranges["stress"][0])
    final_end = _bound_index(ranges["final"][1])
    if not stress_start < validation_start < discovery_start < final_end:
        raise ValueError(
            "auto ranges must satisfy: stress-start < validation-start < "
            "discovery-start < auto-end"
        )


def _auto_configuration(args, profile, grid, base_tune, base_description, resolved_end):
    ranges = _auto_ranges(args, resolved_end)
    _validate_auto_ranges(ranges)
    return {
        "profile": profile,
        "parameter_grid": {key: list(values) for key, values in grid.items()},
        "base_source": base_description,
        "base_tune": base_tune,
        "tests_per_cycle": args.auto_tests,
        "validation_top": args.auto_validation_top,
        "stress_top": args.auto_stress_top,
        "final_top": args.auto_final_top,
        "hall_size": args.auto_hall_size,
        "importance_target": args.auto_importance_target,
        "seed": args.seed,
        "minimum_trades": args.min_trades,
        "maximum_allowed_drawdown": args.max_drawdown,
        "ranges": ranges,
    }


def _auto_feature_configuration(args):
    """Settings for the version-2 search engine, stored outside legacy config."""
    return {
        "advanced_min_candidates": int(args.auto_advanced_min_candidates),
        "halving_rungs": int(args.auto_halving_rungs),
        "halving_keep": float(args.auto_halving_keep),
        "surrogate_min_samples": int(args.auto_surrogate_min_samples),
        "surrogate_pool_multiplier": int(args.auto_surrogate_pool),
        "surrogate_trees": int(args.auto_surrogate_trees),
        "walk_forward_folds": int(args.auto_walk_forward_folds),
        "walk_forward_top": int(args.auto_walk_forward_top),
        "walk_forward_stability_penalty": float(
            args.auto_walk_forward_stability_penalty
        ),
    }


def _auto_bootstrap(args, grid):
    """Use an existing standard-optimizer winner to warm-start a new campaign."""
    if getattr(args, "base_source", "config") != "config":
        return None
    path = Path(getattr(args, "base_params", "outputs/optimize/best_params.json"))
    if not path.is_file():
        return None
    saved = load_ma_strategy_tune(path)
    compatible = {key: saved[key] for key in grid if key in saved}
    if not compatible:
        return None
    return {
        "source": str(path),
        "compatible_parameter_count": len(compatible),
        "params": compatible,
    }


def _load_json(path, default=None):
    path = Path(path)
    if not path.is_file():
        return default
    with path.open(encoding="utf-8") as file:
        return json.load(file)


def _candidate_signature(params, keys):
    return tuple(params[key] for key in keys)


def _load_auto_seen(output_dir, keys):
    seen = set()
    paths = sorted((Path(output_dir) / "cycles").glob("cycle_*/*_candidates.json"))
    for index, path in enumerate(paths, start=1):
        for candidate in _read_candidate_plan(path):
            params = candidate.get("params", {})
            if all(key in params for key in keys):
                seen.add(_candidate_signature(params, keys))
        _show_loading_progress("Loading previous candidates", index, len(paths))
    return seen


def _load_discovery_history(output_dir, keys):
    """Load a compact persistent surrogate history and increment it as needed."""
    output_dir = Path(output_dir)
    cache_path = output_dir / "surrogate_history_cache.json.gz"
    cycles_dir = Path(output_dir) / "cycles"
    plan_paths = sorted(cycles_dir.glob("cycle_*/discovery_candidates.json"))
    cached_history, cached_cycle = _read_surrogate_history_cache(cache_path, keys)
    if cached_history is None:
        history = []
        cached_cycle = 0
        selected_paths = _surrogate_bootstrap_paths(
            plan_paths, SURROGATE_CACHE_BOOTSTRAP_CYCLES
        )
        if len(selected_paths) < len(plan_paths):
            print(
                f"Building compact surrogate cache from {len(selected_paths):,}/"
                f"{len(plan_paths):,} time-balanced/recent cycles."
            )
    else:
        history = cached_history
        selected_paths = [
            path for path in plan_paths
            if _cycle_number_from_path(path) > cached_cycle
        ]
        print(
            f"Loaded compact surrogate cache: {len(history):,} samples through "
            f"cycle {cached_cycle:,}."
        )

    for index, plan_path in enumerate(selected_paths, start=1):
        results_path = _resolve_csv_path(plan_path.with_name("discovery_results.csv"))
        candidates = _read_candidate_plan(plan_path)
        for record in _read_auto_stage_records(results_path, candidates):
            score = _finite_number(record.get("objective_score"))
            if score is not None and all(key in record["params"] for key in keys):
                history.append(record)
        _show_loading_progress("Loading surrogate history", index, len(selected_paths))

    latest_cycle = max(
        [cached_cycle] + [_cycle_number_from_path(path) for path in selected_paths]
    )
    if selected_paths or cached_history is None:
        history = _write_surrogate_history_cache(
            cache_path, history, keys, latest_cycle
        )
    return history


def _cycle_number_from_path(path):
    try:
        return int(Path(path).parent.name.removeprefix("cycle_"))
    except ValueError:
        return 0


def _evenly_spaced_items(items, limit):
    items = list(items)
    limit = max(1, int(limit))
    if len(items) <= limit:
        return items
    if limit == 1:
        return [items[-1]]
    indices = {
        round(position * (len(items) - 1) / (limit - 1))
        for position in range(limit)
    }
    return [items[index] for index in sorted(indices)]


def _surrogate_bootstrap_paths(paths, limit):
    """Blend whole-history coverage with extra weight on recent refinements."""
    paths = list(paths)
    limit = max(1, int(limit))
    if len(paths) <= limit:
        return paths
    broad_count = max(1, limit // 2)
    broad = _evenly_spaced_items(paths, broad_count)
    selected = {path: None for path in broad}
    for path in reversed(paths):
        selected[path] = None
        if len(selected) >= limit:
            break
    return sorted(selected)


def _read_surrogate_history_cache(path, keys):
    path = Path(path)
    if not path.is_file():
        return None, 0
    try:
        with gzip.open(path, "rt", encoding="utf-8") as cache_file:
            payload = json.load(cache_file)
        if (
            payload.get("version") != SURROGATE_CACHE_VERSION
            or tuple(payload.get("parameter_keys", ())) != tuple(keys)
        ):
            return None, 0
        rows = payload.get("rows", [])
        history = [
            {
                "candidate_id": row[0],
                "objective_score": row[1],
                "params": dict(zip(keys, row[2:])),
            }
            for row in rows
        ]
        return history, int(payload.get("latest_cycle", 0) or 0)
    except (OSError, EOFError, UnicodeError, ValueError, json.JSONDecodeError):
        return None, 0


def _write_surrogate_history_cache(path, history, keys, latest_cycle):
    usable = [
        record for record in history
        if _finite_number(record.get("objective_score")) is not None
        and all(key in record.get("params", {}) for key in keys)
    ]
    selected = _representative_surrogate_history(
        usable, SURROGATE_MAX_TRAINING_SAMPLES
    )
    payload = {
        "version": SURROGATE_CACHE_VERSION,
        "latest_cycle": int(latest_cycle),
        "parameter_keys": list(keys),
        "sample_count": len(selected),
        "sample_method": "deterministic_score_quantiles",
        "rows": [
            [
                record.get("candidate_id"),
                record["objective_score"],
                *(record["params"][key] for key in keys),
            ]
            for record in selected
        ],
    }
    path = Path(path)
    temporary = path.with_suffix(path.suffix + ".tmp")
    try:
        with gzip.open(
            temporary, "wt", encoding="utf-8", compresslevel=6
        ) as cache_file:
            json.dump(_json_safe(payload), cache_file, separators=(",", ":"))
        os.replace(temporary, path)
    except BaseException:
        if temporary.is_file():
            temporary.unlink()
        raise
    return selected


def _surrogate_feature_row(params, keys, grid):
    row = []
    for key in keys:
        value = params[key]
        if isinstance(value, bool):
            row.append(float(value))
        elif isinstance(value, (int, float)):
            row.append(float(value))
        else:
            values = list(grid[key])
            row.append(float(values.index(value)) if value in values else -1.0)
    return row


def _representative_surrogate_history(history, limit):
    """Select deterministic score quantiles while retaining the complete history on disk."""
    history = list(history)
    limit = max(2, int(limit))
    if len(history) <= limit:
        return history
    ranked = sorted(history, key=lambda record: float(record["objective_score"]))
    indices = {
        round(position * (len(ranked) - 1) / (limit - 1))
        for position in range(limit)
    }
    return [ranked[index] for index in sorted(indices)]


def _select_surrogate_candidates(
    candidates, history, count, keys, grid, features, seed,
):
    """Rank a large candidate pool with historical data and retain exploration."""
    candidates = list(candidates)
    usable_history = [
        record for record in history
        if _finite_number(record.get("objective_score")) is not None
    ]
    minimum = features["surrogate_min_samples"]
    if len(usable_history) < minimum or len(candidates) <= count:
        return candidates[:count], {
            "enabled": False,
            "reason": "insufficient_history" if len(usable_history) < minimum else "small_pool",
            "history_samples": len(usable_history),
            "minimum_samples": minimum,
            "candidate_pool": len(candidates),
        }

    training_history = _representative_surrogate_history(
        usable_history, SURROGATE_MAX_TRAINING_SAMPLES
    )
    if len(training_history) < len(usable_history):
        print(
            f"Surrogate training set: {len(training_history):,} representative "
            f"samples from {len(usable_history):,} stored results."
        )
    training_features = [
        _surrogate_feature_row(record["params"], keys, grid)
        for record in training_history
    ]
    targets = [float(record["objective_score"]) for record in training_history]
    model = ExtraTreesSurrogate(
        n_trees=features["surrogate_trees"],
        max_depth=max(6, min(12, round(math.log2(len(training_history) + 1)) + 1)),
        min_leaf=max(3, min(12, len(training_history) // 40)),
        seed=seed,
    ).fit(training_features, targets, progress_label="Training surrogate model")
    predictions = model.predict_mean_std(
        [_surrogate_feature_row(candidate, keys, grid) for candidate in candidates],
        progress_label="Scoring candidate pool",
    )
    ranked_mean = sorted(
        range(len(candidates)), key=lambda index: predictions[index][0], reverse=True
    )
    ranked_uncertainty = sorted(
        range(len(candidates)), key=lambda index: predictions[index][1], reverse=True
    )
    selected_indices = []
    selected_set = set()

    def add(indices, limit):
        for index in indices:
            if len(selected_indices) >= limit:
                break
            if index not in selected_set:
                selected_set.add(index)
                selected_indices.append(index)

    exploit_target = max(1, round(count * 0.55))
    uncertainty_target = min(count, exploit_target + round(count * 0.20))
    add(ranked_mean, exploit_target)
    add(ranked_uncertainty, uncertainty_target)
    remaining = [index for index in range(len(candidates)) if index not in selected_set]
    random.Random(seed + 7919).shuffle(remaining)
    add(remaining, count)
    selected = [candidates[index] for index in selected_indices[:count]]
    selected_predictions = [predictions[index] for index in selected_indices[:count]]
    return selected, {
        "enabled": True,
        "model": "dependency_free_extra_trees",
        "history_samples": len(usable_history),
        "training_samples": len(training_history),
        "training_sample_method": "deterministic_score_quantiles",
        "candidate_pool": len(candidates),
        "selected_candidates": len(selected),
        "trees": features["surrogate_trees"],
        "predicted_score_min": min(item[0] for item in selected_predictions),
        "predicted_score_max": max(item[0] for item in selected_predictions),
        "mean_uncertainty": statistics.fmean(item[1] for item in selected_predictions),
        "selection_mix": {"exploitation": 0.55, "uncertainty": 0.20, "random": 0.25},
    }


def _promote_candidates(records, keep_count):
    ranked = sorted(
        records,
        key=lambda record: (
            _finite_number(record.get("objective_score"))
            if _finite_number(record.get("objective_score")) is not None
            else -math.inf
        ),
        reverse=True,
    )
    return [
        {"candidate_id": record["candidate_id"], "params": record["params"]}
        for record in ranked[:keep_count]
        if _finite_number(record.get("objective_score")) is not None
    ]


def _expanding_discovery_ranges(range_start, range_end, rung_count):
    start_index = _bound_index(range_start)
    end_index = _bound_index(range_end)
    width = end_index - start_index
    if width <= rung_count + 1:
        return []
    ranges = []
    for rung in range(1, rung_count + 1):
        fraction = rung / (rung_count + 1)
        rung_start = end_index - max(1, round(width * fraction))
        ranges.append((rung_start, end_index))
    return ranges


def _walk_forward_ranges(ranges, fold_count):
    start_index = _bound_index(ranges["stress"][0])
    end_index = _bound_index(ranges["discovery"][0])
    width = end_index - start_index
    fold_count = min(max(0, fold_count), max(0, width))
    if fold_count < 2:
        return []
    boundaries = [
        start_index + round(width * index / fold_count)
        for index in range(fold_count + 1)
    ]
    return [
        (boundaries[index], boundaries[index + 1])
        for index in range(fold_count)
        if boundaries[index] < boundaries[index + 1]
    ]


def _count_auto_evaluations(output_dir):
    completed = 0
    cycles_dir = Path(output_dir) / "cycles"
    paths = list(cycles_dir.glob("cycle_*/*_results.csv"))
    paths.extend(
        path for path in cycles_dir.glob("cycle_*/*_results.csv.gz")
        if not path.with_suffix("").is_file()
    )
    for index, path in enumerate(paths, start=1):
        with _open_csv_text(path) as csv_file:
            reader = csv.DictReader(csv_file)
            completed += sum(1 for _ in reader)
        _show_loading_progress("Counting legacy results", index, len(paths))
    return completed


def _resolve_csv_path(path):
    """Prefer a writable/plain CSV and otherwise use its completed gzip archive."""
    path = Path(path)
    if path.is_file():
        return path
    compressed = path.with_suffix(path.suffix + ".gz")
    return compressed if compressed.is_file() else path


def _open_csv_text(path):
    path = Path(path)
    if path.suffix == ".gz":
        return gzip.open(path, "rt", newline="", encoding="utf-8")
    return path.open(newline="", encoding="utf-8")


def _count_csv_rows(path):
    path = _resolve_csv_path(path)
    if not path.is_file():
        return 0
    with _open_csv_text(path) as csv_file:
        return sum(1 for _ in csv.DictReader(csv_file))


def _reconcile_auto_evaluations(output_dir, state):
    """Trust persisted totals and only reconcile the active stage after a crash."""
    saved_total = state.get("total_evaluations")
    if saved_total is None:
        return _count_auto_evaluations(output_dir)
    cycle_dir = (
        Path(output_dir) / "cycles" / f"cycle_{int(state.get('cycle', 1)):06d}"
    )
    stage = state.get("stage")
    if not stage:
        return int(saved_total)
    actual_stage_rows = _count_csv_rows(cycle_dir / f"{stage}_results.csv")
    saved_stage_rows = int(state.get("stage_completed", 0) or 0)
    recovered_rows = max(0, actual_stage_rows - saved_stage_rows)
    if recovered_rows:
        print(
            f"Resume reconciliation: recovered {recovered_rows:,} completed "
            f"{stage} result(s) written after the last state flush."
        )
    return int(saved_total) + recovered_rows


def _write_candidate_plan(path, cycle, stage, candidates, source=None):
    """Write a compact, backwards-compatible candidate plan.

    Parameter names are stored once instead of once per candidate.  A stage that
    reuses an unchanged plan can point at its source rather than duplicating it.
    """
    path = Path(path)
    if source is not None:
        payload = {
            "version": 2,
            "cycle": cycle,
            "stage": stage,
            "candidate_count": len(candidates),
            "source": os.path.relpath(Path(source), path.parent),
        }
    else:
        parameter_keys = list(candidates[0].get("params", {})) if candidates else []
        payload = {
            "version": 2,
            "cycle": cycle,
            "stage": stage,
            "candidate_count": len(candidates),
            "parameter_keys": parameter_keys,
            "candidate_rows": [
                [candidate["candidate_id"], *(
                    candidate.get("params", {}).get(key) for key in parameter_keys
                )]
                for candidate in candidates
            ],
        }
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(_json_safe(payload), file, separators=(",", ":"), ensure_ascii=False)
        file.write("\n")
    os.replace(temporary, path)


def _read_candidate_plan(path):
    payload = _load_json(path, {}) or {}
    if payload.get("source"):
        source_path = Path(path).parent / payload["source"]
        return _read_candidate_plan(source_path.resolve())
    if "candidate_rows" in payload:
        keys = payload.get("parameter_keys", [])
        return [
            {
                "candidate_id": row[0],
                "params": dict(zip(keys, row[1:])),
            }
            for row in payload["candidate_rows"]
        ]
    return payload.get("candidates", [])


def _compact_auto_candidate_plans(output_dir, show_progress=False):
    """Upgrade verbose legacy plans and replace exact rung-1 copies in place."""
    rewritten = 0
    bytes_before = 0
    bytes_after = 0
    cycles_dir = Path(output_dir) / "cycles"
    paths = sorted(cycles_dir.glob("cycle_*/*_candidates.json"))
    for index, path in enumerate(paths, start=1):
        payload = _load_json(path, {}) or {}
        needs_upgrade = payload.get("version") != 2
        source_path = path.parent / "discovery_pool_candidates.json"
        use_source = path.name == "discovery_rung_01_candidates.json" and source_path.is_file()
        if not needs_upgrade and (not use_source or payload.get("source")):
            if show_progress:
                _show_loading_progress("Checking candidate plans", index, len(paths))
            continue
        candidates = _read_candidate_plan(path)
        if use_source and candidates != _read_candidate_plan(source_path):
            use_source = False
        old_size = path.stat().st_size
        _write_candidate_plan(
            path,
            payload.get("cycle"),
            payload.get("stage", path.stem.removesuffix("_candidates")),
            candidates,
            source=source_path if use_source else None,
        )
        rewritten += 1
        bytes_before += old_size
        bytes_after += path.stat().st_size
        if show_progress:
            _show_loading_progress("Checking candidate plans", index, len(paths))
    return rewritten, max(0, bytes_before - bytes_after)


def _compress_completed_cycle_results(cycle_dir):
    """Losslessly archive completed CSVs while keeping their columns easy to scan."""
    compressed_count = 0
    reclaimed = 0
    known_fields = _auto_stage_fieldnames(())
    known_set = set(known_fields)
    for path in sorted(Path(cycle_dir).glob("*_results.csv")):
        destination = path.with_suffix(path.suffix + ".gz")
        original_size = path.stat().st_size
        if destination.is_file():
            path.unlink()
            compressed_count += 1
            reclaimed += original_size
            continue
        temporary = destination.with_suffix(destination.suffix + ".tmp")
        try:
            with path.open(newline="", encoding="utf-8") as source:
                reader = csv.DictReader(source)
                existing_fields = reader.fieldnames or []
                parameter_fields = [
                    field for field in existing_fields if field not in known_set
                ]
                fieldnames = [
                    field for field in known_fields if field != "error" and field in existing_fields
                ]
                fieldnames.extend(parameter_fields)
                if "error" in existing_fields:
                    fieldnames.append("error")
                with gzip.open(
                    temporary, "wt", newline="", encoding="utf-8", compresslevel=6
                ) as target:
                    writer = csv.DictWriter(target, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(reader)
            os.replace(temporary, destination)
            compressed_size = destination.stat().st_size
            path.unlink()
            compressed_count += 1
            reclaimed += max(0, original_size - compressed_size)
        except BaseException:
            if temporary.is_file():
                temporary.unlink()
            raise
    return compressed_count, reclaimed


def _cleanup_completed_cycle(cycle_dir):
    """Compress and prune exactly one completed cycle."""
    removed = 0
    reclaimed = 0
    cycle_dir = Path(cycle_dir)
    try:
        compressed, cycle_reclaimed = _compress_completed_cycle_results(cycle_dir)
    except (OSError, UnicodeError, csv.Error) as error:
        # Storage optimization is optional; a locked or unusual legacy CSV must
        # not prevent the actual optimization campaign from continuing.
        print(f"Storage warning: kept legacy files in {cycle_dir.name}: {error}")
        compressed = 0
    else:
        reclaimed += cycle_reclaimed
    keep = {"discovery_candidates.json", "discovery_pool_candidates.json"}
    for path in cycle_dir.glob("*_candidates.json"):
        if path.name not in keep:
            reclaimed += path.stat().st_size
            path.unlink()
            removed += 1
    checkpoints_dir = cycle_dir / "checkpoints"
    if checkpoints_dir.is_dir():
        for checkpoint_dir in checkpoints_dir.iterdir():
            if not checkpoint_dir.is_dir():
                continue
            for name in ("best_params.json", "checkpoint.json"):
                path = checkpoint_dir / name
                if path.is_file():
                    reclaimed += path.stat().st_size
                    path.unlink()
                    removed += 1
            if not any(checkpoint_dir.iterdir()):
                checkpoint_dir.rmdir()
        if not any(checkpoints_dir.iterdir()):
            checkpoints_dir.rmdir()
    return removed, compressed, reclaimed


def _cleanup_completed_auto_cycles(output_dir, current_cycle, show_progress=False):
    """Upgrade every older cycle once when a campaign process starts."""
    removed = compressed = reclaimed = 0
    cycles_dir = Path(output_dir) / "cycles"
    cycle_dirs = sorted(cycles_dir.glob("cycle_*"))
    eligible = []
    for cycle_dir in cycle_dirs:
        try:
            cycle_number = int(cycle_dir.name.removeprefix("cycle_"))
        except ValueError:
            continue
        if cycle_number >= int(current_cycle):
            continue
        eligible.append(cycle_dir)
    for index, cycle_dir in enumerate(eligible, start=1):
        cycle_removed, cycle_compressed, cycle_reclaimed = _cleanup_completed_cycle(
            cycle_dir
        )
        removed += cycle_removed
        compressed += cycle_compressed
        reclaimed += cycle_reclaimed
        if show_progress:
            _show_loading_progress("Preparing stored cycles", index, len(eligible))
    return removed, compressed, reclaimed


def _safe_cleanup_completed_auto_cycles(output_dir, current_cycle, show_progress=False):
    """Keep optional storage cleanup from blocking a compatible resume."""
    try:
        return _cleanup_completed_auto_cycles(
            output_dir, current_cycle, show_progress=show_progress
        )
    except OSError as error:
        print(f"Storage warning: cleanup was deferred: {error}")
        return 0, 0, 0


def _safe_cleanup_completed_cycle(cycle_dir):
    """Clean only the cycle that just completed; older gzip files are untouched."""
    try:
        return _cleanup_completed_cycle(cycle_dir)
    except OSError as error:
        print(f"Storage warning: cleanup was deferred for {Path(cycle_dir).name}: {error}")
        return 0, 0, 0


def _auto_stage_fieldnames(keys):
    metrics = [column for column in RESULT_COLUMNS if column != "score"]
    return [
        "candidate_id", "cycle", "stage", "range_start", "range_end",
        "objective_score", "time_normalized_score", "score", *metrics,
        "profit_per_trade", "range_candles", "duration_s", *keys, "error",
    ]


def _upgrade_auto_stage_csv(path, keys):
    """Add time-normalized columns before appending to a version-1 checkpoint."""
    path = Path(path)
    if not path.is_file() or path.stat().st_size == 0:
        return
    with path.open(newline="", encoding="utf-8") as csv_file:
        reader = csv.DictReader(csv_file)
        existing_fields = reader.fieldnames or []
        target_fields = _auto_stage_fieldnames(keys)
        if existing_fields == target_fields:
            return
        rows = list(reader)
    for row in rows:
        range_candles = _range_candle_count(row["range_start"], row["range_end"])
        row["range_candles"] = range_candles
        row["time_normalized_score"] = _time_normalized_score(
            row.get("objective_score"), range_candles
        )
    _write_rows_atomic(path, _auto_stage_fieldnames(keys), rows)


def _auto_result_row(keys, candidate_id, cycle, stage, range_start, range_end,
                     params, result, duration, objective_score, error,
                     time_normalized_score=None, range_candles=None):
    row = {
        "candidate_id": candidate_id,
        "cycle": cycle,
        "stage": stage,
        "range_start": range_start,
        "range_end": range_end,
        "duration_s": round(duration, 4),
        "objective_score": objective_score,
        "time_normalized_score": time_normalized_score,
        "range_candles": range_candles,
        "error": error,
    }
    row.update({key: params[key] for key in keys})
    if result:
        row.update({key: result.get(key) for key in RESULT_COLUMNS})
        closed_trades = int(result.get("closed_trades", 0) or 0)
        realized_profit = result.get("realized_profit")
        if realized_profit is None:
            realized_profit = result.get("total_profit", 0)
        row["profit_per_trade"] = (
            float(realized_profit or 0) / closed_trades if closed_trades else None
        )
    return row


def _read_auto_stage_records(path, candidates):
    path = _resolve_csv_path(path)
    if not path.is_file():
        return []
    candidate_map = {candidate["candidate_id"]: candidate for candidate in candidates}
    records = {}
    with _open_csv_text(path) as csv_file:
        reader = csv.DictReader(csv_file)
        if "candidate_id" not in (reader.fieldnames or ()):
            raise ValueError(f"invalid auto checkpoint: {path}")
        for row in reader:
            candidate = candidate_map.get(row["candidate_id"])
            if candidate is None:
                continue
            result = {key: _parse_metric(row.get(key)) for key in RESULT_COLUMNS}
            raw_score = _parse_metric(row.get("objective_score"))
            range_candles = _parse_metric(row.get("range_candles"))
            if range_candles is None:
                try:
                    range_candles = _range_candle_count(
                        row.get("range_start"), row.get("range_end")
                    )
                except (TypeError, ValueError):
                    range_candles = None
            normalized_score = _parse_metric(row.get("time_normalized_score"))
            if normalized_score is None:
                normalized_score = _time_normalized_score(raw_score, range_candles)
            records[row["candidate_id"]] = {
                "candidate_id": row["candidate_id"],
                "params": candidate["params"],
                "result": result,
                "objective_score": raw_score,
                "time_normalized_score": normalized_score,
                "range_candles": range_candles,
                "duration": float(row.get("duration_s") or 0),
                "error": row.get("error") or None,
            }
    return [
        records[candidate["candidate_id"]]
        for candidate in candidates
        if candidate["candidate_id"] in records
    ]


def _auto_stage_best(records):
    records = list(records)

    def rank(record):
        objective = _record_comparable_score(record)
        if objective is not None:
            return 1, objective
        score = _finite_number((record.get("result") or {}).get("score"))
        return 0, score if score is not None else -math.inf

    return max(records, key=rank) if records else None


def _write_auto_stage_checkpoint(
    cycle_dir, stage, records, base_tune, completed, total, state
):
    """Keep a best_params file beside every resumable auto-stage checkpoint."""
    best = _auto_stage_best(records)
    if best is None:
        return None
    effective_params = {**base_tune, **best["params"]}
    checkpoint_dir = Path(cycle_dir) / "checkpoints" / stage
    checkpoint_dir.mkdir(parents=True, exist_ok=True)
    _write_json(checkpoint_dir / "best_params.json", effective_params)
    _write_json(Path(cycle_dir) / "best_params.json", effective_params)
    checkpoint = {
        "cycle": state["cycle"],
        "stage": stage,
        "completed": completed,
        "total": total,
        "best_candidate_id": best["candidate_id"],
        "best_objective_score": best.get("objective_score"),
        "best_time_normalized_score": best.get("time_normalized_score"),
        "best_params": effective_params,
        "updated_at": _timestamp_now(),
    }
    _write_json(checkpoint_dir / "checkpoint.json", checkpoint)
    state.update({
        "checkpoint_best_stage": stage,
        "checkpoint_best_candidate_id": best["candidate_id"],
        "checkpoint_best_params": effective_params,
    })
    return best


def _clear_auto_stage_checkpoint(cycle_dir, stage):
    """Remove an ephemeral stage checkpoint after its result CSV is complete."""
    checkpoint_dir = Path(cycle_dir) / "checkpoints" / stage
    for name in ("best_params.json", "checkpoint.json"):
        path = checkpoint_dir / name
        if path.is_file():
            path.unlink()
    if checkpoint_dir.is_dir() and not any(checkpoint_dir.iterdir()):
        checkpoint_dir.rmdir()
    parent = checkpoint_dir.parent
    if parent.is_dir() and not any(parent.iterdir()):
        parent.rmdir()


def _run_auto_stage(
    args,
    cycle,
    stage,
    candidates,
    range_start,
    range_end,
    base_tune,
    cycle_dir,
    state,
    state_path,
    min_trades_override=None,
):
    """Evaluate one auto stage and persist every yielded result for exact resume."""
    keys = tuple(state["config"]["parameter_grid"])
    results_path = Path(cycle_dir) / f"{stage}_results.csv"
    _upgrade_auto_stage_csv(results_path, keys)
    existing = _read_auto_stage_records(results_path, candidates)
    records = {record["candidate_id"]: record for record in existing}
    pending = [
        candidate for candidate in candidates
        if candidate["candidate_id"] not in records
    ]
    total = len(candidates)
    print(
        f"Cycle {cycle} | {stage}: {len(existing):,}/{total:,} complete | "
        f"range {_format_range_bound(range_start)} -> "
        f"{_format_range_bound(range_end)}"
    )
    if not pending:
        _write_auto_stage_checkpoint(
            cycle_dir, stage, existing, base_tune, len(existing), total, state
        )
        _clear_auto_stage_checkpoint(cycle_dir, stage)
        return existing, False

    fieldnames = _auto_stage_fieldnames(keys)
    append = results_path.is_file() and results_path.stat().st_size > 0
    workers = min(max(1, args.workers), len(pending))
    batch_size = args.batch_size or max(32, workers * 8)
    chunksize = args.chunksize or max(1, batch_size // (workers * 4))
    started = time.perf_counter()
    pool = None
    pool_terminated = False
    interrupted = False
    effective_min_trades = (
        args.min_trades if min_trades_override is None else min_trades_override
    )
    range_candles = _range_candle_count(range_start, range_end)

    with results_path.open("a" if append else "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        if not append:
            writer.writeheader()
        if workers > 1:
            pool = multiprocessing.Pool(
                workers,
                initializer=_init_worker,
                initargs=(
                    _parse_bound(range_start), _parse_bound(range_end), base_tune, True,
                ),
            )
        else:
            _init_worker(
                _parse_bound(range_start), _parse_bound(range_end), base_tune
            )

        tasks = [
            (candidate["candidate_id"], candidate["params"])
            for candidate in pending
        ]
        if pool is not None:
            evaluated = pool.imap_unordered(_evaluate_task, tasks, chunksize=chunksize)
        else:
            evaluated = (
                _evaluate_candidate(
                    candidate_id,
                    params,
                    base_tune,
                    _parse_bound(range_start),
                    _parse_bound(range_end),
                )
                for candidate_id, params in tasks
            )

        try:
            for candidate_id, params, result, duration, error in evaluated:
                objective_score = None
                normalized_score = None
                if not error:
                    closed_trades = int(result.get("closed_trades", 0) or 0)
                    if closed_trades > 0:
                        value = _objective_score(
                            result,
                            min_trades=effective_min_trades,
                            max_drawdown=args.max_drawdown,
                        )
                        objective_score = value if math.isfinite(value) else None
                        normalized_score = _time_normalized_score(
                            objective_score, range_candles
                        )
                writer.writerow(_auto_result_row(
                    keys, candidate_id, cycle, stage, range_start, range_end,
                    params, result, duration, objective_score, error,
                    time_normalized_score=normalized_score,
                    range_candles=range_candles,
                ))
                csv_file.flush()
                records[candidate_id] = {
                    "candidate_id": candidate_id,
                    "params": params,
                    "result": result or {},
                    "objective_score": objective_score,
                    "time_normalized_score": normalized_score,
                    "range_candles": range_candles,
                    "duration": duration,
                    "error": error,
                }
                state["total_evaluations"] += 1
                state["stage_completed"] = len(records)
                current_best = _auto_stage_best(records.values())
                if current_best and current_best["candidate_id"] == candidate_id:
                    _write_auto_stage_checkpoint(
                        cycle_dir, stage, records.values(), base_tune,
                        len(records), total, state,
                    )
                if args.log_every and (
                    len(records) % args.log_every == 0 or len(records) == total
                ):
                    best_score = max(
                        (_finite_number(record["objective_score"]) for record in records.values()),
                        default=None,
                        key=lambda value: -math.inf if value is None else value,
                    )
                    elapsed = time.perf_counter() - started
                    shown = "n/a" if best_score is None else f"{best_score:.4f}"
                    current_shown = (
                        "n/a" if objective_score is None else f"{objective_score:.4f}"
                    )
                    print(
                        f"  [{len(records):,}/{total:,}] best={shown} "
                        f"score={current_shown} elapsed={elapsed:.1f}s"
                    )
                if len(records) % max(1, min(25, args.log_every or 25)) == 0:
                    _write_auto_stage_checkpoint(
                        cycle_dir, stage, records.values(), base_tune,
                        len(records), total, state,
                    )
                    state["updated_at"] = _timestamp_now()
                    _write_json(state_path, state)
        except KeyboardInterrupt:
            interrupted = True
            state.update({
                "status": "interrupted",
                "stage": stage,
                "stage_completed": len(records),
                "updated_at": _timestamp_now(),
            })
            _write_auto_stage_checkpoint(
                cycle_dir, stage, records.values(), base_tune,
                len(records), total, state,
            )
            _write_json(state_path, state)
            if pool is not None:
                pool.terminate()
                pool_terminated = True
            print(
                f"\nAuto mode stopped during {stage} after {len(records):,}/{total:,} "
                "stage tests. Checkpoint saved; use --auto --resume."
            )
        finally:
            if pool is not None:
                if not pool_terminated:
                    pool.close()
                pool.join()

    ordered = [
        records[candidate["candidate_id"]]
        for candidate in candidates
        if candidate["candidate_id"] in records
    ]
    _write_auto_stage_checkpoint(
        cycle_dir, stage, ordered, base_tune, len(ordered), total, state
    )
    if not interrupted:
        _clear_auto_stage_checkpoint(cycle_dir, stage)
    return ordered, interrupted


def _aggregate_walk_forward_records(fold_records, candidates, stability_penalty):
    record_maps = {
        fold_name: {record["candidate_id"]: record for record in records}
        for fold_name, records in fold_records.items()
    }
    aggregated = []
    for candidate in candidates:
        candidate_id = candidate["candidate_id"]
        records = [
            record_maps[fold_name].get(candidate_id)
            for fold_name in sorted(record_maps)
        ]
        if not records or any(record is None for record in records):
            continue
        raw_scores = [_finite_number(record.get("objective_score")) for record in records]
        scores = [_record_comparable_score(record) for record in records]
        qualified = all(score is not None for score in scores)
        valid_scores = [score for score in scores if score is not None]
        if qualified:
            mean_score = statistics.fmean(valid_scores)
            median_score = statistics.median(valid_scores)
            worst_score = min(valid_scores)
            score_std = statistics.pstdev(valid_scores) if len(valid_scores) > 1 else 0.0
            composite = (
                0.50 * median_score
                + 0.25 * mean_score
                + 0.25 * worst_score
                - stability_penalty * score_std
            )
        else:
            mean_score = median_score = worst_score = score_std = None
            composite = -math.inf

        result = {}
        for metric in RESULT_COLUMNS:
            values = [
                _finite_number((record.get("result") or {}).get(metric))
                for record in records
            ]
            numeric = [value for value in values if value is not None]
            if not numeric:
                result[metric] = None
            elif metric in ("closed_trades", "wins", "losses", "liquidations"):
                result[metric] = sum(numeric)
            elif metric == "maximum_drawdown":
                result[metric] = max(numeric, key=abs)
            else:
                result[metric] = statistics.median(numeric)
        result["score"] = composite if math.isfinite(composite) else None
        aggregated.append({
            "candidate_id": candidate_id,
            "params": candidate["params"],
            "result": result,
            "objective_score": composite if math.isfinite(composite) else None,
            "time_normalized_score": composite if math.isfinite(composite) else None,
            "range_candles": sum(record.get("range_candles", 0) or 0 for record in records),
            "duration": sum(record.get("duration", 0) for record in records),
            "error": None if qualified else "candidate failed one or more walk-forward folds",
            "walk_forward": {
                "fold_raw_scores": raw_scores,
                "fold_time_normalized_scores": scores,
                "mean_score": mean_score,
                "median_score": median_score,
                "worst_score": worst_score,
                "score_std": score_std,
                "stability_penalty": stability_penalty,
            },
        })
    return aggregated


def _run_walk_forward_stage(
    args, cycle, candidates, folds, base_tune, cycle_dir, state, state_path,
):
    """Evaluate fixed candidates on disjoint chronological folds with exact resume."""
    fold_records = {}
    for fold_index, (range_start, range_end) in enumerate(folds, start=1):
        fold_name = f"walk_forward_fold_{fold_index:02d}"
        state.update({
            "stage": fold_name,
            "stage_completed": 0,
            "updated_at": _timestamp_now(),
        })
        _write_json(state_path, state)
        records, interrupted = _run_auto_stage(
            args, cycle, fold_name, candidates, range_start, range_end,
            base_tune, cycle_dir, state, state_path,
            min_trades_override=(
                0 if not args.min_trades else max(
                    1, math.ceil(args.min_trades / max(1, len(folds)))
                )
            ),
        )
        fold_records[fold_name] = records
        if interrupted:
            return [], True

    aggregated = _aggregate_walk_forward_records(
        fold_records,
        candidates,
        state["optimizer_features"]["walk_forward_stability_penalty"],
    )
    summary = {
        "cycle": cycle,
        "stage": "walk_forward",
        "folds": [
            {"fold": index, "start": start, "end": end}
            for index, (start, end) in enumerate(folds, start=1)
        ],
        "candidate_count": len(candidates),
        "ranking_formula": (
            "0.50*median + 0.25*mean + 0.25*worst "
            "- stability_penalty*population_stddev"
        ),
        "records": aggregated,
    }
    _write_json(Path(cycle_dir) / "walk_forward_summary.json", summary)
    state.update({
        "stage": "walk_forward",
        "stage_completed": len(aggregated),
        "updated_at": _timestamp_now(),
    })
    _write_auto_stage_checkpoint(
        cycle_dir, "walk_forward", aggregated, base_tune,
        len(aggregated), len(candidates), state,
    )
    _clear_auto_stage_checkpoint(cycle_dir, "walk_forward")
    _write_json(state_path, state)
    return aggregated, False


def _write_rows_atomic(path, fieldnames, rows):
    path = Path(path)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    _replace_with_retry(temporary, path)


def _flatten_hall_record(record, keys, rank):
    row = {
        "rank": rank,
        "cycle": record["cycle"],
        "candidate_id": record["candidate_id"],
        "robust_score": record["robust_score"],
        "worst_stage_percentile": record.get("worst_stage_percentile"),
    }
    row.update({key: record["params"].get(key) for key in keys})
    metric_names = (
        "score", "time_normalized_score", "total_profit", "total_profit_percent", "closed_trades",
        "win_rate", "maximum_drawdown", "profit_factor", "expectancy_percent",
        "calmar_ratio", "liquidations",
    )
    for stage in AUTO_STAGE_ORDER:
        metrics = record.get("stage_metrics", {}).get(stage, {})
        for metric in metric_names:
            row[f"{stage}_{metric}"] = metrics.get(metric)
    return row


def _save_auto_workbook(output_dir, hall_rows, importance_rows):
    if not hall_rows:
        return None
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return None

    output_path = Path(output_dir) / "auto_report.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(hall_rows).to_excel(writer, sheet_name="Hall of Fame", index=False)
        pd.DataFrame(importance_rows).to_excel(
            writer, sheet_name="Parameter Importance", index=False
        )
    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_index, cells in enumerate(worksheet.columns, start=1):
            width = min(38, max(10, max(len(str(cell.value or "")) for cell in cells[:200]) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        if worksheet.max_row >= 2:
            table = Table(displayName=f"AutoOptimizerTable{index}", ref=worksheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True,
                showFirstColumn=False, showLastColumn=False,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for metric in ("robust_score", "weight", "effect"):
            column = headers.get(metric)
            if column and worksheet.max_row >= 3:
                letter = get_column_letter(column)
                worksheet.conditional_formatting.add(
                    f"{letter}2:{letter}{worksheet.max_row}",
                    ColorScaleRule(
                        start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B",
                    ),
                )
    workbook.save(output_path)
    return output_path


def _write_auto_reports(output_dir, hall, importance, state, keys, excel_enabled=True):
    output_dir = Path(output_dir)
    ranked_hall = sorted(hall, key=lambda record: record["robust_score"], reverse=True)
    _write_json(output_dir / "hall_of_fame.json", ranked_hall)
    _write_json(output_dir / "parameter_importance.json", importance)
    if ranked_hall:
        _write_json(output_dir / "best_params.json", ranked_hall[0]["effective_params"])
    elif state.get("checkpoint_best_params"):
        # Before the first full finalist exists, expose the best provisional
        # checkpoint so even an early interruption has a usable best_params.json.
        _write_json(output_dir / "best_params.json", state["checkpoint_best_params"])
    hall_rows = [
        _flatten_hall_record(record, keys, rank)
        for rank, record in enumerate(ranked_hall, start=1)
    ]
    importance_rows = [
        {"rank": rank, "parameter": key, **item}
        for rank, (key, item) in enumerate(
            sorted(importance.items(), key=lambda pair: pair[1]["weight"], reverse=True),
            start=1,
        )
    ]
    if hall_rows:
        _write_rows_atomic(output_dir / "hall_of_fame.csv", list(hall_rows[0]), hall_rows)
    if importance_rows:
        _write_rows_atomic(
            output_dir / "parameter_importance.csv",
            list(importance_rows[0]),
            importance_rows,
        )
    workbook = (
        _save_auto_workbook(output_dir, hall_rows, importance_rows)
        if excel_enabled else None
    )
    summary = {
        "mode": "auto",
        "optimizer_version": state.get("version", 1),
        "status": state["status"],
        "cycles_completed": state["cycles_completed"],
        "current_cycle": state["cycle"],
        "current_stage": state["stage"],
        "total_evaluations": state["total_evaluations"],
        "hall_of_fame_size": len(ranked_hall),
        "best_robust_score": ranked_hall[0]["robust_score"] if ranked_hall else None,
        "best_params": ranked_hall[0]["effective_params"] if ranked_hall else None,
        "optimizer_features": state.get("optimizer_features"),
        "advanced_from_cycle": state.get("advanced_from_cycle"),
        "bootstrap": state.get("bootstrap"),
        "excel_report": str(workbook) if workbook else None,
        "updated_at": state["updated_at"],
    }
    _write_json(output_dir / "auto_summary.json", summary)


def _merge_hall_of_fame(hall, finalists, cycle, keys, base_tune, limit):
    by_signature = {
        _candidate_signature(record["params"], keys): record for record in hall
    }
    for finalist in finalists:
        if not math.isfinite(finalist["robust_score"]):
            continue
        record = {
            **finalist,
            "cycle": cycle,
            "effective_params": {**base_tune, **finalist["params"]},
        }
        signature = _candidate_signature(record["params"], keys)
        previous = by_signature.get(signature)
        if previous is None or record["robust_score"] > previous["robust_score"]:
            by_signature[signature] = record
    return sorted(
        by_signature.values(), key=lambda record: record["robust_score"], reverse=True
    )[:limit]


def run_auto_optimization(args, grid=None):
    """Run an unlimited, staged, importance-guided optimization campaign."""
    profile = getattr(args, "profile", None) or "full"
    grid = PARAMETER_PROFILES[profile] if grid is None else grid
    keys = tuple(grid)
    base_tune, base_description = _load_base_tune(args)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    state_path = output_dir / "auto_state.json"
    resolved_end = _latest_market_end() if args.auto_end == "latest" else args.auto_end
    config = _auto_configuration(
        args, profile, grid, base_tune, base_description, resolved_end
    )
    optimizer_features = _auto_feature_configuration(args)
    bootstrap = _auto_bootstrap(args, grid)
    resume_existing = bool(args.resume or state_path.is_file())

    if resume_existing:
        state = _load_json(state_path)
        if state is None:
            raise FileNotFoundError(
                f"auto checkpoint not found: {state_path}. Start without --resume first."
            )
        if state.get("version", 1) not in LEGACY_AUTO_STATE_VERSIONS:
            raise ValueError("auto checkpoint version is not compatible")
        if state.get("config") != config:
            raise ValueError(
                "auto checkpoint settings differ from this command; use the original "
                "profile, ranges, test counts, base parameters, and constraints"
            )
        saved_features = state.get("optimizer_features")
        if saved_features is not None and saved_features != optimizer_features:
            raise ValueError(
                "auto search-engine settings differ from this checkpoint; use the "
                "original halving, surrogate, and walk-forward options"
            )
        if saved_features is None:
            # Version-1 campaigns may already be halfway through a stage. Finish
            # that exact plan first and activate the new engine next cycle.
            current_cycle_dir = (
                output_dir / "cycles" / f"cycle_{int(state['cycle']):06d}"
            )
            state["optimizer_features"] = optimizer_features
            state["advanced_from_cycle"] = int(state["cycle"]) + int(
                current_cycle_dir.exists()
            )
            state["migrated_from_version"] = state.get("version", 1)
            state["version"] = AUTO_STATE_VERSION
        hall = _load_json(output_dir / "hall_of_fame.json", []) or []
        importance = _load_json(output_dir / "parameter_importance.json", {}) or {}
        state.update({
            "status": "running",
            "total_evaluations": _reconcile_auto_evaluations(output_dir, state),
            "updated_at": _timestamp_now(),
        })
        print(
            f"Resuming auto cycle {state['cycle']} at {state['stage']} | "
            f"{state['total_evaluations']:,} total evaluations"
        )
        if not args.resume:
            print("Existing auto checkpoint detected; resume was selected automatically.")
    else:
        if state_path.exists():
            raise FileExistsError(
                f"an auto campaign already exists in {output_dir}; use --auto --resume "
                "or choose another --output-dir"
            )
        equal_weight = 1.0 / max(1, len(keys))
        importance = {
            key: {"weight": equal_weight, "effect": 0.0, "groups": 0, "samples": 0}
            for key in keys
        }
        hall = []
        state = {
            "version": AUTO_STATE_VERSION,
            "status": "running",
            "cycle": 1,
            "cycles_completed": 0,
            "stage": "discovery",
            "stage_completed": 0,
            "total_evaluations": 0,
            "importance_cycle": 0,
            "created_at": _timestamp_now(),
            "updated_at": _timestamp_now(),
            "config": config,
            "optimizer_features": optimizer_features,
            "advanced_from_cycle": 1,
            "bootstrap": bootstrap,
        }
    removed_files, compressed_csvs, removed_bytes = _safe_cleanup_completed_auto_cycles(
        output_dir, state["cycle"], show_progress=resume_existing
    )
    try:
        compacted_plans, compacted_bytes = _compact_auto_candidate_plans(
            output_dir, show_progress=resume_existing
        )
    except OSError as error:
        print(f"Storage warning: legacy candidate plans were kept unchanged: {error}")
        compacted_plans = compacted_bytes = 0
    reclaimed_bytes = removed_bytes + compacted_bytes
    if removed_files or compressed_csvs or compacted_plans:
        print(
            f"Storage cleanup: removed {removed_files:,} redundant files, "
            f"compressed {compressed_csvs:,} result CSVs, "
            f"compacted {compacted_plans:,} candidate plans | "
            f"reclaimed {reclaimed_bytes / (1024 * 1024):.1f} MiB"
        )
    _write_json(state_path, state)

    ranges = config["ranges"]
    print(
        f"Mode: auto | {args.auto_tests:,} discovery tests/cycle | "
        f"workers: {max(1, args.workers)}"
    )
    if args.auto_tests >= optimizer_features["advanced_min_candidates"]:
        print(
            f"Funnel: {args.auto_tests:,} -> {args.auto_validation_top} -> "
            f"{args.auto_stress_top} -> WF {args.auto_walk_forward_top} -> "
            f"{args.auto_final_top} | profile: {profile}"
        )
    else:
        print(
            f"Funnel (legacy-small-run): {args.auto_tests:,} -> "
            f"{args.auto_validation_top} -> {args.auto_stress_top} -> "
            f"{args.auto_final_top} | profile: {profile}"
        )
    print(f"Latest market end: {resolved_end}")
    if not resume_existing and bootstrap:
        print(
            f"Warm start: {bootstrap['compatible_parameter_count']} compatible "
            f"parameters loaded from {bootstrap['source']}"
        )

    cycle_limit = max(0, int(args.auto_cycles))
    # Disk history is loaded once per process. Subsequent cycles extend these
    # caches with only their new plans/results instead of rescanning everything.
    seen_cache = None
    history_cache = None
    try:
        while cycle_limit == 0 or state["cycles_completed"] < cycle_limit:
            cycle = int(state["cycle"])
            cycle_dir = output_dir / "cycles" / f"cycle_{cycle:06d}"
            cycle_dir.mkdir(parents=True, exist_ok=True)
            continuation_parent = hall[0] if hall else None
            continuation_base = (
                continuation_parent.get("effective_params", continuation_parent["params"])
                if continuation_parent
                else (state.get("bootstrap") or {}).get("params", base_tune)
            )
            training_parent = {
                "source": "hall_of_fame" if continuation_parent else "base_configuration",
                "candidate_id": (
                    continuation_parent.get("candidate_id") if continuation_parent else None
                ),
                "robust_score": (
                    continuation_parent.get("robust_score") if continuation_parent else None
                ),
                "params": continuation_base,
            }
            state["training_parent"] = training_parent
            _write_json(cycle_dir / "training_parent.json", training_parent)
            advanced = (
                args.auto_tests >= optimizer_features["advanced_min_candidates"]
                and cycle >= int(state.get("advanced_from_cycle", 1))
            )
            source_plan_path = cycle_dir / (
                "discovery_pool_candidates.json" if advanced
                else "discovery_candidates.json"
            )
            if source_plan_path.is_file():
                discovery_pool = _read_candidate_plan(source_plan_path)
            else:
                if seen_cache is None:
                    seen_cache = _load_auto_seen(output_dir, keys)
                else:
                    print(
                        f"Using in-memory candidate history: "
                        f"{len(seen_cache):,} tested configurations."
                    )
                elite_records = [
                    {
                        "params": {key: record["params"][key] for key in keys},
                        "result": record.get("stage_metrics", {}).get("final", {}),
                    }
                    for record in hall
                ]
                generator = SmartCandidateGenerator(
                    grid,
                    seed=args.seed + cycle - 1,
                    baseline_params=continuation_base,
                    continuous_refinement=bool(elite_records),
                    parameter_importance={
                        key: item.get("weight", 1.0) for key, item in importance.items()
                    },
                    refinement_round=max(1, cycle - 1),
                )
                generator.seen.update(seen_cache)
                if advanced:
                    if history_cache is None:
                        history_cache = _load_discovery_history(output_dir, keys)
                    else:
                        print(
                            f"Using in-memory surrogate history: "
                            f"{len(history_cache):,} stored results."
                        )
                    history = history_cache
                else:
                    history = []
                surrogate_ready = (
                    len(history) >= optimizer_features["surrogate_min_samples"]
                )
                pool_target = args.auto_tests * (
                    optimizer_features["surrogate_pool_multiplier"]
                    if advanced and surrogate_ready else 1
                )
                generated_pool = (
                    generator.generate_auto(
                        pool_target,
                        elites=elite_records,
                        progress_label="Generating candidate pool",
                    )
                    if elite_records else generator.generate(
                        pool_target, progress_label="Generating candidate pool"
                    )
                )
                if len(generated_pool) < pool_target:
                    print(
                        f"\nCandidate pool reached {len(generated_pool):,}/{pool_target:,} "
                        "unique valid configurations; continuing with the available pool."
                    )
                generated, surrogate_metadata = _select_surrogate_candidates(
                    generated_pool,
                    history,
                    args.auto_tests,
                    keys,
                    grid,
                    optimizer_features,
                    args.seed + cycle - 1,
                )
                seen_cache.update(
                    _candidate_signature(params, keys) for params in generated
                )
                discovery_pool = [
                    {
                        "candidate_id": f"c{cycle:06d}-{index:06d}",
                        "params": params,
                    }
                    for index, params in enumerate(generated, start=1)
                ]
                _write_candidate_plan(
                    source_plan_path,
                    cycle,
                    "discovery_pool" if advanced else "discovery",
                    discovery_pool,
                )
                if advanced:
                    surrogate_metadata.update({
                        "cycle": cycle,
                        "created_at": _timestamp_now(),
                        "uses_previous_discovery_results": True,
                    })
                    _write_json(cycle_dir / "surrogate_search.json", surrogate_metadata)
            if not discovery_pool:
                state.update({"status": "exhausted", "updated_at": _timestamp_now()})
                _write_json(state_path, state)
                print("Auto search space is exhausted; no unique discovery candidates remain.")
                break

            stage_results = {}
            discovery_candidates = discovery_pool
            if advanced and optimizer_features["halving_rungs"]:
                rung_ranges = _expanding_discovery_ranges(
                    *ranges["discovery"], optimizer_features["halving_rungs"]
                )
                previous_records = None
                for rung_index, rung_range in enumerate(rung_ranges, start=1):
                    rung_stage = f"discovery_rung_{rung_index:02d}"
                    rung_plan_path = cycle_dir / f"{rung_stage}_candidates.json"
                    if rung_plan_path.is_file():
                        rung_candidates = _read_candidate_plan(rung_plan_path)
                    elif rung_index == 1:
                        rung_candidates = discovery_pool
                        _write_candidate_plan(
                            rung_plan_path, cycle, rung_stage, rung_candidates,
                            source=source_plan_path,
                        )
                    else:
                        keep_count = max(
                            args.auto_validation_top,
                            math.ceil(
                                len(discovery_candidates)
                                * optimizer_features["halving_keep"]
                            ),
                        )
                        rung_candidates = _promote_candidates(
                            previous_records, keep_count
                        )
                        _write_candidate_plan(
                            rung_plan_path, cycle, rung_stage, rung_candidates
                        )
                    discovery_candidates = rung_candidates
                    state.update({
                        "status": "running", "stage": rung_stage,
                        "stage_completed": 0, "updated_at": _timestamp_now(),
                    })
                    _write_json(state_path, state)
                    previous_records, interrupted = _run_auto_stage(
                        args, cycle, rung_stage, rung_candidates,
                        *rung_range, base_tune, cycle_dir, state, state_path,
                        min_trades_override=(
                            0 if not args.min_trades else max(
                                1,
                                math.ceil(
                                    args.min_trades
                                    * (rung_range[1] - rung_range[0])
                                    / max(
                                        1,
                                        _bound_index(ranges["discovery"][1])
                                        - _bound_index(ranges["discovery"][0]),
                                    )
                                ),
                            )
                        ),
                    )
                    if interrupted:
                        _write_auto_reports(
                            output_dir, hall, importance, state, keys,
                            excel_enabled=bool(args.excel_top),
                        )
                        return hall[0] if hall else None

                discovery_plan_path = cycle_dir / "discovery_candidates.json"
                if discovery_plan_path.is_file():
                    discovery_candidates = _read_candidate_plan(discovery_plan_path)
                else:
                    keep_count = max(
                        args.auto_validation_top,
                        math.ceil(
                            len(discovery_candidates)
                            * optimizer_features["halving_keep"]
                        ),
                    )
                    discovery_candidates = _promote_candidates(
                        previous_records, keep_count
                    )
                    _write_candidate_plan(
                        discovery_plan_path, cycle, "discovery", discovery_candidates
                    )

            discovery_plan_path = cycle_dir / "discovery_candidates.json"
            if advanced and not discovery_plan_path.is_file():
                _write_candidate_plan(
                    discovery_plan_path, cycle, "discovery", discovery_candidates
                )

            state.update({
                "status": "running", "stage": "discovery", "stage_completed": 0,
                "updated_at": _timestamp_now(),
            })
            _write_json(state_path, state)
            discovery_records, interrupted = _run_auto_stage(
                args, cycle, "discovery", discovery_candidates,
                *ranges["discovery"], base_tune, cycle_dir, state, state_path,
            )
            stage_results["discovery"] = discovery_records
            if interrupted:
                _write_auto_reports(
                    output_dir, hall, importance, state, keys,
                    excel_enabled=bool(args.excel_top),
                )
                return hall[0] if hall else None
            if history_cache is not None:
                history_cache.extend(discovery_records)
                history_cache = _write_surrogate_history_cache(
                    output_dir / "surrogate_history_cache.json.gz",
                    history_cache,
                    keys,
                    cycle,
                )

            if int(state.get("importance_cycle", 0)) < cycle:
                learned = _learn_parameter_importance(
                    discovery_records, keys, target=args.auto_importance_target
                )
                importance = _smooth_parameter_importance(importance, learned)
                state["importance_cycle"] = cycle
                state["updated_at"] = _timestamp_now()
                _write_json(output_dir / "parameter_importance.json", importance)
                _write_json(state_path, state)

            stage_plan = [
                ("validation", args.auto_validation_top),
                ("stress", args.auto_stress_top),
            ]
            if not advanced:
                stage_plan.append(("final", args.auto_final_top))
            for stage, keep_count in stage_plan:
                plan_path = cycle_dir / f"{stage}_candidates.json"
                if plan_path.is_file():
                    candidates = _read_candidate_plan(plan_path)
                else:
                    ranked = _combine_auto_stage_records(stage_results)
                    candidates = [
                        {
                            "candidate_id": record["candidate_id"],
                            "params": record["params"],
                        }
                        for record in ranked[:keep_count]
                        if math.isfinite(record["robust_score"])
                    ]
                    _write_candidate_plan(plan_path, cycle, stage, candidates)
                state.update({
                    "stage": stage, "stage_completed": 0, "updated_at": _timestamp_now(),
                })
                _write_json(state_path, state)
                records, interrupted = _run_auto_stage(
                    args, cycle, stage, candidates, *ranges[stage], base_tune,
                    cycle_dir, state, state_path,
                )
                stage_results[stage] = records
                if interrupted:
                    _write_auto_reports(
                        output_dir, hall, importance, state, keys,
                        excel_enabled=bool(args.excel_top),
                    )
                    return hall[0] if hall else None

            if advanced:
                walk_folds = _walk_forward_ranges(
                    ranges, optimizer_features["walk_forward_folds"]
                )
                if walk_folds:
                    walk_plan_path = cycle_dir / "walk_forward_candidates.json"
                    if walk_plan_path.is_file():
                        walk_candidates = _read_candidate_plan(walk_plan_path)
                    else:
                        ranked = _combine_auto_stage_records(stage_results)
                        walk_candidates = [
                            {
                                "candidate_id": record["candidate_id"],
                                "params": record["params"],
                            }
                            for record in ranked[
                                :optimizer_features["walk_forward_top"]
                            ]
                            if math.isfinite(record["robust_score"])
                        ]
                        _write_candidate_plan(
                            walk_plan_path, cycle, "walk_forward", walk_candidates
                        )
                    walk_records, interrupted = _run_walk_forward_stage(
                        args, cycle, walk_candidates, walk_folds, base_tune,
                        cycle_dir, state, state_path,
                    )
                    stage_results["walk_forward"] = walk_records
                    if interrupted:
                        _write_auto_reports(
                            output_dir, hall, importance, state, keys,
                            excel_enabled=bool(args.excel_top),
                        )
                        return hall[0] if hall else None

                final_plan_path = cycle_dir / "final_candidates.json"
                if final_plan_path.is_file():
                    final_candidates = _read_candidate_plan(final_plan_path)
                else:
                    ranked = _combine_auto_stage_records(stage_results)
                    final_candidates = [
                        {
                            "candidate_id": record["candidate_id"],
                            "params": record["params"],
                        }
                        for record in ranked[:args.auto_final_top]
                        if math.isfinite(record["robust_score"])
                    ]
                    _write_candidate_plan(
                        final_plan_path, cycle, "final", final_candidates
                    )
                state.update({
                    "stage": "final", "stage_completed": 0,
                    "updated_at": _timestamp_now(),
                })
                _write_json(state_path, state)
                final_records, interrupted = _run_auto_stage(
                    args, cycle, "final", final_candidates, *ranges["final"],
                    base_tune, cycle_dir, state, state_path,
                )
                stage_results["final"] = final_records
                if interrupted:
                    _write_auto_reports(
                        output_dir, hall, importance, state, keys,
                        excel_enabled=bool(args.excel_top),
                    )
                    return hall[0] if hall else None

            finalists = _combine_auto_stage_records(stage_results)
            hall = _merge_hall_of_fame(
                hall, finalists, cycle, keys, base_tune, args.auto_hall_size
            )
            state.update({
                "cycles_completed": state["cycles_completed"] + 1,
                "cycle": cycle + 1,
                "stage": "discovery",
                "stage_completed": 0,
                "status": "running",
                "updated_at": _timestamp_now(),
            })
            _write_json(state_path, state)
            _write_auto_reports(
                output_dir, hall, importance, state, keys,
                excel_enabled=bool(args.excel_top),
            )
            _safe_cleanup_completed_cycle(cycle_dir)
            best_text = (
                f"{hall[0]['robust_score']:.4f}" if hall else "no qualified finalist"
            )
            important = sorted(
                importance.items(), key=lambda pair: pair[1]["weight"], reverse=True
            )[:5]
            print(
                f"Cycle {cycle} complete | Hall of Fame: {len(hall)} | best: {best_text}"
            )
            print(
                "Most influential parameters: "
                + ", ".join(f"{key} ({item['weight']:.1%})" for key, item in important)
            )
    except KeyboardInterrupt:
        state.update({"status": "interrupted", "updated_at": _timestamp_now()})
        _write_json(state_path, state)
        _write_auto_reports(
            output_dir, hall, importance, state, keys,
            excel_enabled=bool(args.excel_top),
        )
        print("\nAuto mode stopped safely. Use --auto --resume to continue.")
        return hall[0] if hall else None

    state.update({"status": "completed", "updated_at": _timestamp_now()})
    _write_json(state_path, state)
    _write_auto_reports(
        output_dir, hall, importance, state, keys,
        excel_enabled=bool(args.excel_top),
    )
    print(
        f"Auto campaign stopped after {state['cycles_completed']} completed cycle(s). "
        f"Resume with --auto --resume."
    )
    return hall[0] if hall else None


def run_optimization(args, grid=None):
    profile = getattr(args, "profile", None) or "focused"
    grid = PARAMETER_PROFILES[profile] if grid is None else grid
    keys = tuple(grid)
    base_tune, base_description = _load_base_tune(args)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    results_path = output_dir / "optimization_results.csv"
    fieldnames = _result_fieldnames(keys)

    if args.mode == "grid":
        requested_tests = grid_size(grid)
        candidate_source = None
    else:
        requested_tests = min(args.tests, grid_size(grid))
        candidate_source = None

    workers = max(1, args.workers)
    batch_size = args.batch_size or max(32, workers * 8)
    chunksize = args.chunksize or max(1, batch_size // (workers * 4))
    start = _parse_bound(args.start)
    end = _parse_bound(args.end)
    started = time.perf_counter()
    completed = 0
    failed = 0
    next_index = 1
    best = None
    ranked = []
    min_trades = getattr(args, "min_trades", 0)
    max_drawdown = getattr(args, "max_drawdown", None)

    def objective(result):
        return _objective_score(result, min_trades=min_trades, max_drawdown=max_drawdown)

    resume = bool(getattr(args, "resume", False))
    resume_signatures = set()
    if resume and results_path.is_file():
        ranked = _read_resume_records(results_path, grid, base_tune)
        completed = len(ranked)
        resume_signatures = {
            tuple(record["params"][key] for key in keys) for record in ranked
        }
        if ranked:
            next_index = max(record["index"] for record in ranked) + 1
        ranked.sort(key=lambda item: objective(item["result"]), reverse=True)
        best = next(
            (record for record in ranked if math.isfinite(objective(record["result"]))),
            None,
        )
        del ranked[max(
            getattr(args, "top_n", 20), args.elite_size,
            getattr(args, "validation_top", 20),
        ):]
        print(f"Resuming from {completed:,} completed candidates in {results_path}")

    seen_signatures = resume_signatures
    if args.mode == "grid":
        candidate_source = (
            candidate for candidate in iter_grid_candidates(grid)
            if tuple(candidate[key] for key in keys) not in seen_signatures
            and is_valid_candidate({**base_tune, **candidate})
        )

    run_metadata = {
        "profile": profile,
        "optimized_parameters": list(keys),
        "base_source": base_description,
        "minimum_trades": min_trades,
        "maximum_allowed_drawdown": max_drawdown,
        "resumed": resume,
    }

    print(f"Mode: {args.mode} | tests: {requested_tests:,} | workers: {workers}")
    print(f"Profile: {profile} ({len(keys)} parameters) | base: {base_description}")
    print(f"Range: {start} -> {end}")

    append_results = resume and results_path.is_file() and completed > 0
    with results_path.open("a" if append_results else "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        if not append_results:
            writer.writeheader()

        pool = None
        if workers > 1:
            pool = multiprocessing.Pool(
                workers,
                initializer=_init_worker,
                initargs=(start, end, base_tune, True),
            )
        else:
            # Warm cached market data once in the parent for serial searches.
            _init_worker(start, end, base_tune)

        def evaluate(batch):
            tasks = [(offset, candidate) for offset, candidate in batch]
            if pool is not None:
                return pool.imap_unordered(_evaluate_task, tasks, chunksize=chunksize)
            return (
                _evaluate_candidate(index, params, base_tune, start, end)
                for index, params in tasks
            )

        def consume(batch):
            nonlocal completed, failed, best, ranked
            for index, params, result, duration, error in evaluate(batch):
                completed += 1
                if error:
                    failed += 1
                    print(f"[{completed}/{requested_tests}] test {index} failed: {error}")
                    continue
                effective_params = {**base_tune, **params}
                result_objective = objective(result)
                writer.writerow(_result_row(
                    keys, index, effective_params, result, duration,
                    objective_score=(result_objective if math.isfinite(result_objective) else None),
                ))
                record = {
                    "index": index,
                    "params": effective_params,
                    "result": result,
                    "duration": duration,
                }
                ranked.append(record)
                ranked.sort(key=lambda item: objective(item["result"]), reverse=True)
                del ranked[max(
                    getattr(args, "top_n", 20), args.elite_size,
                    getattr(args, "validation_top", 20),
                ):]
                if math.isfinite(objective(result)) and (
                    best is None or objective(result) > objective(best["result"])
                ):
                    best = record
                if args.log_every and (completed % args.log_every == 0 or completed == requested_tests):
                    best_score = _score(best["result"]) if best else -math.inf
                    current_score = _score(result)
                    elapsed = time.perf_counter() - started
                    print(
                        f"[{completed:,}/{requested_tests:,}] best_score={best_score:.4f} "
                        f"score={current_score:.4f} elapsed={elapsed:.1f}s"
                    )
            csv_file.flush()
            _write_best_files(
                output_dir, best, args.mode, requested_tests, completed,
                time.perf_counter() - started, args.seed, run_metadata,
            )

        interrupted = False
        pool_terminated = False
        try:
            if args.mode == "grid":
                while True:
                    candidates = list(itertools.islice(candidate_source, batch_size))
                    if not candidates:
                        break
                    batch = list(enumerate(candidates, start=next_index))
                    next_index += len(batch)
                    consume(batch)
            else:
                generator = SmartCandidateGenerator(
                    grid, seed=args.seed, baseline_params=base_tune,
                )
                generator.seen.update(seen_signatures)
                while completed < requested_tests:
                    count = min(batch_size, requested_tests - completed)
                    elites = [
                        record for record in ranked[:args.elite_size]
                        if math.isfinite(objective(record["result"]))
                    ]
                    progress = completed / max(1, requested_tests)
                    candidates = generator.generate(count, elites=elites, progress=progress)
                    if not candidates:
                        break
                    batch = list(enumerate(candidates, start=next_index))
                    next_index += len(batch)
                    consume(batch)
        except KeyboardInterrupt:
            interrupted = True
            csv_file.flush()
            _write_best_files(
                output_dir, best, args.mode, requested_tests, completed,
                time.perf_counter() - started, args.seed,
                {**run_metadata, "interrupted": True},
            )
            if pool is not None:
                pool.terminate()
                pool_terminated = True
            print(
                f"\nStopped by user after {completed:,} completed tests. "
                "Checkpoint saved; use --resume to continue."
            )
        finally:
            if pool is not None:
                if not pool_terminated:
                    pool.close()
                pool.join()

    if interrupted:
        return best

    elapsed = time.perf_counter() - started
    training_best = best
    validated_best, validation_records = _validate_top_candidates(
        ranked, args, workers, chunksize,
    )
    if validation_records:
        _write_json(output_dir / "validation_results.json", validation_records)
        run_metadata.update({
            "validation_range": [args.validation_start, args.validation_end],
            "validation_candidates": len(validation_records),
            "overfit_penalty": args.overfit_penalty,
        })
        if training_best:
            _write_json(output_dir / "best_training_params.json", training_best["params"])
        if validated_best and math.isfinite(validated_best["robust_score"]):
            best = {
                "index": validated_best["index"],
                "params": validated_best["params"],
                "result": validated_best["result"],
                "duration": validated_best["duration"],
            }
            run_metadata.update({
                "best_robust_score": validated_best["robust_score"],
                "best_training_metrics": validated_best["training_result"],
            })
            print(
                f"Validated {len(validation_records)} finalists | "
                f"best robust score={validated_best['robust_score']:.4f}"
            )
        else:
            run_metadata["validation_warning"] = "no finalist passed validation constraints"
            print("Validation warning: no finalist passed the requested constraints")
    elapsed = time.perf_counter() - started
    excel_top = max(0, int(getattr(args, "excel_top", 5000)))
    workbook_path = (
        _save_optimizer_workbook(
            results_path,
            output_dir / "optimization_results.xlsx",
            keys,
            max_rows=excel_top,
        )
        if excel_top else None
    )
    if workbook_path is not None:
        run_metadata["excel_report"] = str(workbook_path)
    _write_best_files(
        output_dir, best, args.mode, requested_tests, completed, elapsed,
        args.seed, run_metadata,
    )
    top_n = max(1, int(getattr(args, "top_n", 20)))
    _write_json(output_dir / "top_results.json", [
        {"rank": rank, **record}
        for rank, record in enumerate(ranked[:top_n], start=1)
    ])
    print(f"Finished {completed:,} tests ({failed} failed) in {elapsed:.1f}s")
    print(f"Results: {results_path}")
    if workbook_path is not None:
        print(f"Excel report: {workbook_path}")
    if best:
        print(f"Best score: {_score(best['result']):.4f}")
        print(f"Best params: {output_dir / 'best_params.json'}")
    return best


class _OptimizerHelpFormatter(
    argparse.ArgumentDefaultsHelpFormatter,
    argparse.RawDescriptionHelpFormatter,
):
    """Keep command examples readable while still showing option defaults."""

    def _get_help_string(self, action):
        help_text = action.help
        if (
            "%(default)" not in help_text
            and action.default not in (None, False, argparse.SUPPRESS)
        ):
            help_text += " (default: %(default)s)"
        return help_text


def build_parser():
    parser = argparse.ArgumentParser(
        prog="optimize.py",
        formatter_class=_OptimizerHelpFormatter,
        description="""Search for robust ma_strategy parameters.

Choose one search path:
  smart  Budgeted adaptive search (recommended for normal experiments).
  grid   Every combination in a profile; usually only practical for tiny grids.
  auto   Continuous model-guided search with successive halving, walk-forward
         validation, stress tests, and full-history finalists. Existing state is
         resumed automatically. It runs until Ctrl+C unless --auto-cycles is set.

Dates are inclusive at START and exclusive at END. A candle index may be used
instead of a YYYY-MM-DD date.""",
        epilog="""recommended examples:
  Inspect profiles and estimate a run without starting it:
    python optimize.py --list-profiles
    python optimize.py --mode smart --profile focused --tests 5000 --dry-run

  Start a fast adaptive search from strategy_config.py:
    python optimize.py --mode smart --profile focused --base-source config `
      --tests 5000 -w 8 --output-dir outputs/optimize/focused_run

  Train on one period and validate the best 30 candidates on unseen data:
    python optimize.py --mode smart --profile signal --tests 10000 -w 8 `
      --start 2023-01-01 --end 2025-01-01 `
      --validation-start 2025-01-01 --validation-end 2026-01-01 `
      --validation-top 30 --min-trades 50 --max-drawdown 35 `
      --output-dir outputs/optimize/validated_signal

  Refine an existing winner, then resume the same interrupted run:
    python optimize.py --mode smart --profile exit --base-source best `
      --base-params outputs/optimize/best_params.json --tests 5000 -w 8 `
      --output-dir outputs/optimize/refine_exit
    python optimize.py --mode smart --profile exit --base-source best `
      --base-params outputs/optimize/best_params.json --tests 5000 -w 8 `
      --output-dir outputs/optimize/refine_exit --resume

  Run two auto cycles (omit --auto-cycles to run until Ctrl+C):
    python optimize.py --auto --auto-cycles 2 -w 16 `
      --output-dir outputs/optimize/auto_two_cycles

  Resume the auto campaign with exactly the same settings (--resume is optional
  when the checkpoint already exists):
    python optimize.py --auto --auto-cycles 2 -w 16 `
      --output-dir outputs/optimize/auto_two_cycles --resume

Tips:
  * Start with --dry-run. Full built-in grids can contain enormous combinations.
  * Use a new --output-dir for a new experiment; use --resume only for the same run.
  * Plain --auto detects and resumes a compatible checkpoint in --output-dir.
  * A new campaign warm-starts from a compatible existing --base-params winner.
  * For trustworthy selection, keep validation/stress data outside the search range.
  * Raw score is preserved; cross-range comparisons use a candle-count annualized score.
  * Auto mode defaults to profile=full; non-auto mode defaults to profile=focused.""",
    )

    search = parser.add_argument_group("search mode and parameter scope")
    search.add_argument(
        "--auto", action="store_true",
        help="use the resumable staged auto campaign (overrides --mode)",
    )
    search.add_argument(
        "--mode", choices=("smart", "grid"), default="smart",
        help="non-auto search algorithm",
    )
    search.add_argument(
        "--tests", type=int, default=5000, metavar="N",
        help="candidate budget in smart mode; ignored by grid and auto",
    )
    search.add_argument(
        "--profile", choices=tuple(PARAMETER_PROFILES), default=None,
        help="parameter group (default: full in auto mode, focused otherwise)",
    )
    search.add_argument(
        "--base-source", choices=("config", "best", "file"), default="config",
        help="fixed/base values come from strategy_config.py or --base-params JSON",
    )
    search.add_argument(
        "--base-params", default=os.path.join("outputs", "optimize", "best_params.json"),
        metavar="PATH", help="JSON read when --base-source is best or file",
    )

    execution = parser.add_argument_group("execution and reproducibility")
    execution.add_argument(
        "-w", "--workers", type=int, default=min(8, os.cpu_count() or 1), metavar="N",
        help="parallel worker processes",
    )
    execution.add_argument(
        "--batch-size", type=int, default=0, metavar="N",
        help="candidates evaluated before adapting/checkpointing; 0 selects automatically",
    )
    execution.add_argument(
        "--chunksize", type=int, default=0, metavar="N",
        help="tasks sent to each worker at once; 0 selects automatically",
    )
    execution.add_argument(
        "--elite-size", type=int, default=20, metavar="N",
        help="top candidates that guide smart search",
    )
    execution.add_argument(
        "--seed", type=int, default=42, metavar="N",
        help="random seed for reproducible smart/auto candidate generation",
    )

    ranges = parser.add_argument_group("standard search ranges and robustness")
    ranges.add_argument(
        "--start", default="2025-01-01", metavar="DATE|INDEX",
        help="inclusive training start",
    )
    ranges.add_argument(
        "--end", default="2026-02-23", metavar="DATE|INDEX",
        help="exclusive training end",
    )
    ranges.add_argument(
        "--validation-start", metavar="DATE|INDEX",
        help="inclusive out-of-sample start; requires --validation-end",
    )
    ranges.add_argument(
        "--validation-end", metavar="DATE|INDEX",
        help="exclusive out-of-sample end; requires --validation-start",
    )
    ranges.add_argument(
        "--validation-top", type=int, default=20, metavar="N",
        help="training finalists re-tested out-of-sample",
    )
    ranges.add_argument(
        "--overfit-penalty", type=float, default=0.25, metavar="FLOAT",
        help="penalty when training score exceeds validation score",
    )
    ranges.add_argument(
        "--min-trades", type=int, default=0, metavar="N",
        help="disqualify candidates with fewer closed trades (0 disables)",
    )
    ranges.add_argument(
        "--max-drawdown", type=float, metavar="PERCENT",
        help="disqualify candidates above this absolute drawdown percentage",
    )

    output = parser.add_argument_group("output, checkpoints, and planning")
    output.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, metavar="PATH",
                        help="directory for CSV, JSON, checkpoints, and reports")
    output.add_argument("--resume", action="store_true",
                        help="require a compatible checkpoint; auto detects one without this flag")
    output.add_argument("--log-every", type=int, default=10, metavar="N",
                        help="print progress every N completed tests (0 is silent)")
    output.add_argument("--top-n", type=int, default=20, metavar="N",
                        help="ranked candidates saved to top_results.json")
    output.add_argument(
        "--excel-top", type=int, default=5000,
        metavar="N", help="top candidates included in XLSX (0 disables XLSX)",
    )
    output.add_argument("--list-profiles", action="store_true",
                        help="show profile parameter counts/grid sizes and exit")
    output.add_argument("--dry-run", action="store_true",
                        help="print the resolved plan without running backtests")

    auto = parser.add_argument_group("auto campaign (used only with --auto)")
    auto.add_argument(
        "--auto-tests", type=int, default=2000,
        metavar="N",
        help="new discovery candidates generated in every auto cycle",
    )
    auto.add_argument(
        "--auto-validation-top", type=int, default=30,
        metavar="N",
        help="discovery finalists sent to the independent validation range",
    )
    auto.add_argument(
        "--auto-stress-top", type=int, default=10,
        metavar="N",
        help="validation finalists sent to the older stress range",
    )
    auto.add_argument(
        "--auto-final-top", type=int, default=3,
        metavar="N",
        help="stress finalists tested on the complete market range",
    )
    auto.add_argument(
        "--auto-hall-size", type=int, default=20,
        metavar="N",
        help="maximum robust winners retained across all auto cycles",
    )
    auto.add_argument(
        "--auto-cycles", type=int, default=0,
        metavar="N",
        help="stop after N completed cycles (0 runs until Ctrl+C)",
    )
    auto.add_argument(
        "--auto-discovery-start", default="2025-01-01",
        metavar="DATE|INDEX",
        help="start of the recent discovery range",
    )
    auto.add_argument(
        "--auto-validation-start", default="2023-01-01",
        metavar="DATE|INDEX",
        help="start of validation; it ends at auto-discovery-start",
    )
    auto.add_argument(
        "--auto-stress-start", default="2019-01-01",
        metavar="DATE|INDEX",
        help="start of stress testing and the complete final range",
    )
    auto.add_argument(
        "--auto-end", default="latest",
        metavar="DATE|INDEX|latest",
        help="exclusive campaign end or 'latest' to detect the final candle",
    )
    auto.add_argument(
        "--auto-importance-target",
        choices=("objective_score", "total_profit", "total_profit_percent"),
        default="objective_score",
        help="metric used to learn which parameters deserve more mutations",
    )
    auto.add_argument(
        "--auto-advanced-min-candidates", type=int, default=64, metavar="N",
        help="minimum cycle size that activates halving/surrogate/walk-forward",
    )
    auto.add_argument(
        "--auto-halving-rungs", type=int, default=2, metavar="N",
        help="cheap expanding Discovery rungs before full Discovery (0 disables)",
    )
    auto.add_argument(
        "--auto-halving-keep", type=float, default=0.25, metavar="RATIO",
        help="fraction promoted after each cheap Discovery rung",
    )
    auto.add_argument(
        "--auto-surrogate-min-samples", type=int, default=64, metavar="N",
        help="historical full-Discovery samples required before Extra Trees is used",
    )
    auto.add_argument(
        "--auto-surrogate-pool", type=int, default=8, metavar="MULTIPLIER",
        help="unevaluated candidate-pool size relative to --auto-tests",
    )
    auto.add_argument(
        "--auto-surrogate-trees", type=int, default=32, metavar="N",
        help="randomized regression trees in the internal surrogate model",
    )
    auto.add_argument(
        "--auto-walk-forward-folds", type=int, default=3, metavar="N",
        help="disjoint pre-Discovery time folds (0 disables; minimum enabled value is 2)",
    )
    auto.add_argument(
        "--auto-walk-forward-top", type=int, default=10, metavar="N",
        help="Stress finalists evaluated on every walk-forward fold",
    )
    auto.add_argument(
        "--auto-walk-forward-stability-penalty", type=float, default=0.15,
        metavar="FLOAT",
        help="penalty multiplier applied to score variation across folds",
    )
    return parser


def main(argv=None):
    args = build_parser().parse_args(argv)
    args.profile = args.profile or ("full" if args.auto else "focused")
    if args.list_profiles:
        for name, grid in PARAMETER_PROFILES.items():
            print(f"{name}: {len(grid)} parameters | {grid_size(grid):,} grid combinations")
        return
    if args.tests <= 0:
        raise SystemExit("--tests must be greater than zero")
    if args.elite_size <= 0:
        raise SystemExit("--elite-size must be greater than zero")
    if args.validation_top <= 0:
        raise SystemExit("--validation-top must be greater than zero")
    if args.min_trades < 0:
        raise SystemExit("--min-trades cannot be negative")
    if args.max_drawdown is not None and args.max_drawdown <= 0:
        raise SystemExit("--max-drawdown must be greater than zero")
    if args.overfit_penalty < 0:
        raise SystemExit("--overfit-penalty cannot be negative")
    if args.top_n <= 0:
        raise SystemExit("--top-n must be greater than zero")
    if args.excel_top < 0:
        raise SystemExit("--excel-top cannot be negative")
    if args.auto_tests <= 0:
        raise SystemExit("--auto-tests must be greater than zero")
    if min(
        args.auto_validation_top, args.auto_stress_top,
        args.auto_final_top, args.auto_hall_size,
    ) <= 0:
        raise SystemExit("auto funnel and Hall of Fame sizes must be greater than zero")
    if not (
        args.auto_tests >= args.auto_validation_top
        >= args.auto_stress_top >= args.auto_final_top
    ):
        raise SystemExit(
            "auto funnel must satisfy: auto-tests >= auto-validation-top >= "
            "auto-stress-top >= auto-final-top"
        )
    if args.auto_cycles < 0:
        raise SystemExit("--auto-cycles cannot be negative")
    if args.auto_advanced_min_candidates <= 0:
        raise SystemExit("--auto-advanced-min-candidates must be greater than zero")
    if not 0 <= args.auto_halving_rungs <= 4:
        raise SystemExit("--auto-halving-rungs must be between 0 and 4")
    if not 0 < args.auto_halving_keep <= 1:
        raise SystemExit("--auto-halving-keep must be greater than 0 and at most 1")
    if args.auto_surrogate_min_samples < 4:
        raise SystemExit("--auto-surrogate-min-samples must be at least 4")
    if args.auto_surrogate_pool <= 0 or args.auto_surrogate_trees <= 0:
        raise SystemExit("auto surrogate pool and tree counts must be greater than zero")
    if args.auto_walk_forward_folds == 1 or args.auto_walk_forward_folds < 0:
        raise SystemExit("--auto-walk-forward-folds must be 0 or at least 2")
    if args.auto_walk_forward_top <= 0:
        raise SystemExit("--auto-walk-forward-top must be greater than zero")
    if (
        args.auto_tests >= args.auto_advanced_min_candidates
        and not args.auto_stress_top >= args.auto_walk_forward_top >= args.auto_final_top
    ):
        raise SystemExit(
            "advanced auto funnel must satisfy: auto-stress-top >= "
            "auto-walk-forward-top >= auto-final-top"
        )
    if args.auto_walk_forward_stability_penalty < 0:
        raise SystemExit("--auto-walk-forward-stability-penalty cannot be negative")
    if bool(args.validation_start) != bool(args.validation_end):
        raise SystemExit("--validation-start and --validation-end must be used together")
    if args.dry_run:
        selected_grid = PARAMETER_PROFILES[args.profile]
        if args.auto:
            resolved_end = _latest_market_end() if args.auto_end == "latest" else args.auto_end
            ranges = _auto_ranges(args, resolved_end)
            _validate_auto_ranges(ranges)
            print("Mode: auto")
            print(f"Profile: {args.profile} ({len(selected_grid)} parameters)")
            print(
                f"Funnel per cycle: {args.auto_tests:,} -> "
                f"{args.auto_validation_top} -> {args.auto_stress_top} -> "
                f"WF {args.auto_walk_forward_top} x {args.auto_walk_forward_folds} -> "
                f"{args.auto_final_top}"
            )
            print(
                f"Discovery halving: {args.auto_halving_rungs} rung(s), "
                f"keep {args.auto_halving_keep:.0%} per rung"
            )
            print(
                f"Surrogate: {args.auto_surrogate_trees} Extra Trees after "
                f"{args.auto_surrogate_min_samples} historical samples"
            )
            print(f"Cycles: {'unlimited' if args.auto_cycles == 0 else args.auto_cycles}")
            print(f"Workers: {args.workers}")
            for stage in ("discovery", "validation", "stress", "final"):
                print(f"{stage.title()}: {ranges[stage][0]} -> {ranges[stage][1]}")
            for index, fold in enumerate(
                _walk_forward_ranges(ranges, args.auto_walk_forward_folds), start=1
            ):
                print(f"Walk-forward fold {index}: {fold[0]} -> {fold[1]}")
            return
        planned = (
            min(args.tests, grid_size(selected_grid))
            if args.mode == "smart"
            else grid_size(selected_grid)
        )
        print(f"Mode: {args.mode}")
        print(f"Profile: {args.profile} ({len(selected_grid)} parameters)")
        print(f"Planned candidates: {planned:,}")
        print(f"Workers: {args.workers}")
        print(f"Range: {args.start} -> {args.end}")
        return
    multiprocessing.freeze_support()
    if args.auto:
        if args.output_dir == DEFAULT_OUTPUT_DIR:
            args.output_dir = os.path.join(DEFAULT_OUTPUT_DIR, "auto")
        run_auto_optimization(args)
    else:
        run_optimization(args)


if __name__ == "__main__":
    main()
