import pandas as pd
import os
from collections import Counter


class TradeCSVLogger:
    """Lightweight CSV logger.
    - In normal mode it collects rows and writes a CSV on save_csv().
    - In optimize mode (optimize=True) it becomes a no-op to avoid disk I/O
      and reduce per-trade overhead (much faster for grid search).
    """
    COLUMNS = [
        "trade_id",
        "type",
        "open_time",
        "close_time",
        "entry_price",
        "close_price",
        "tactical_balance",
        "balance_before",
        "balance_after",
        "total_assets",
        "amount",
        "profit",
        "profit_percent",
        "pnl_percent",
        "fee_paid",
        "leverage",
        "trade_amount_percent",
        "duration_minutes_total",
        "duration_days",
        "duration_hours",
        "duration_minutes",
        "save_money",
        "profit_percent_per_month",
        "other_open_positions_at_close",
        "reason",
    ]

    def __init__(self, optimize: bool = False, write_excel: bool = True):
        self.optimize = bool(optimize)
        self.write_excel = bool(write_excel)
        if self.optimize:
            # keep only a tiny counter to preserve minimal bookkeeping
            self._count = 0
        else:
            self.rows = []

    def log_trade(
        self,
        trade_id,
        trade_type,
        open_time,
        close_time,
        entry_price,
        close_price,
        tactical_balance,
        total_assets,
        balance_before,
        balance_after,
        margin,
        leverage,
        trade_amount_percent,
        profit,
        profit_percent,
        pnl_percent,
        fee,
        days,
        hours,
        minutes,
        save_money,
        profit_percent_per_month,
        other_open_positions_at_close,
        reason
    ):
        if self.optimize:
            # no per-trade allocations during optimization
            self._count += 1
            return

        self.rows.append({
            "trade_id": trade_id,
            "type": trade_type,
            "open_time": open_time,
            "close_time": close_time,
            "entry_price": entry_price,
            "close_price": close_price,
            "tactical_balance": tactical_balance,
            "total_assets": total_assets,
            "balance_before": balance_before,
            "balance_after": balance_after,
            "amount": margin,
            "leverage": leverage,
            "trade_amount_percent": trade_amount_percent,
            "profit": profit,
            "profit_percent": profit_percent,
            "pnl_percent": pnl_percent,
            "fee_paid": fee,
            "duration_minutes_total": days * 24 * 60 + hours * 60 + minutes,
            "duration_days": days,
            "duration_hours": hours,
            "duration_minutes": minutes,
            "save_money": save_money,
            "profit_percent_per_month": profit_percent_per_month,
            "other_open_positions_at_close": bool(other_open_positions_at_close),
            "reason": reason,
        })

    def save_csv(
        self,
        first_balance,
        final_balance,
        total_profit,
        total_profit_percent,
        total_fee,
        start_time,
        end_time,
        days,
        hours,
        minutes,
        overview_metrics=None,
        file_name: str = os.path.join("outputs", "trades", "data_orders.csv")
    ):
        if self.optimize:
            # do not write any files during optimization
            return {"rows_logged": getattr(self, "_count", 0)}

        df = pd.DataFrame(self.rows, columns=self.COLUMNS)

        summary_row = {
            "trade_id": None,
            "type": "SUMMARY",
            "open_time": start_time,
            "close_time": end_time,
            "entry_price": None,
            "close_price": None,
            "total_assets": final_balance,
            "balance_before": first_balance,
            "balance_after": final_balance,
            "profit": total_profit,
            "profit_percent": total_profit_percent,
            "fee_paid": total_fee,
            "duration_days": days,
            "duration_hours": hours,
            "duration_minutes": minutes
        }
        summary_row_full = {col: summary_row.get(col, None) for col in self.COLUMNS}

        if df.empty:
            df = pd.DataFrame([summary_row_full], columns=self.COLUMNS)
        else:
            summary_frame = pd.DataFrame(
                [summary_row_full], columns=self.COLUMNS, dtype=object
            )
            df = pd.concat(
                [df.astype(object), summary_frame],
                ignore_index=True,
            )
        while True:
            try:
                output_dir = os.path.dirname(file_name)
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                df.to_csv(file_name, index=False, encoding="utf-8")
                if self.write_excel:
                    self._save_colored_excel(df, file_name, overview_metrics)
                break
            except PermissionError:
                answer = input(f"please close: {file_name} after close write ok: ")
                if answer == "ok":
                    print("thanks!")

    def _save_colored_excel(
        self, df: pd.DataFrame, csv_file_name: str, overview_metrics=None
    ):
        """Create a polished multi-sheet workbook alongside the raw CSV."""
        if df.empty:
            return

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except Exception:
            return

        excel_file_name = os.path.splitext(csv_file_name)[0] + ".xlsx"
        summary_df = df[df["type"] == "SUMMARY"].copy()
        trades_df = df[df["type"] != "SUMMARY"].copy()
        trade_ids = trades_df["trade_id"].fillna("").astype(str)
        rsi_df = trades_df[trade_ids.str.startswith("rsi_ma_strategy_")].copy()
        scale_df = trades_df[trade_ids.str.startswith("scale_ma_strategy_")].copy()
        main_df = trades_df[
            ~trade_ids.str.startswith(("rsi_ma_strategy_", "scale_ma_strategy_"))
        ].copy()

        summary_values = summary_df.iloc[0].to_dict() if not summary_df.empty else {}
        if overview_metrics:
            overview_rows = [
                (section, metric, value)
                for section, metrics in overview_metrics.items()
                for metric, value in metrics.items()
            ]
        else:
            overview_rows = [
                ("Run", "Start time", summary_values.get("open_time")),
                ("Run", "End time", summary_values.get("close_time")),
                ("Capital", "Starting balance", summary_values.get("balance_before")),
                ("Capital", "Final balance", summary_values.get("balance_after")),
                ("Capital", "Total profit", summary_values.get("profit")),
                ("Capital", "Total profit %", summary_values.get("profit_percent")),
                ("Capital", "Total fees", summary_values.get("fee_paid")),
                ("Trades", "Closed trades", len(trades_df)),
                ("Trades", "Main trades", len(main_df)),
                ("RSI", "RSI trades", len(rsi_df)),
                ("Scale", "Scale trades", len(scale_df)),
            ]
        overview = pd.DataFrame(overview_rows, columns=["Section", "Metric", "Value"])

        sheets = {
            "Overview": overview,
            "All Trades": trades_df,
            "Main Strategy": main_df,
            "RSI Strategy": rsi_df,
            "Scale Strategy": scale_df,
        }
        with pd.ExcelWriter(excel_file_name, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

        close_times = [str(x) for x in trades_df.get("close_time", [])]
        types = [str(x) for x in trades_df.get("type", [])]
        valid_close_times = [
            ct for ct, t in zip(close_times, types)
            if ct and ct.lower() != "none" and t != "SUMMARY"
        ]
        close_counts = Counter(valid_close_times)
        multi_close_times = {ct for ct, c in close_counts.items() if c >= 2}
        wb = load_workbook(excel_file_name)
        header_fill = PatternFill("solid", fgColor="17365D")
        header_font = Font(color="FFFFFF", bold=True)
        profit_fill = PatternFill("solid", fgColor="E2F0D9")
        loss_fill = PatternFill("solid", fgColor="FCE4D6")
        grouped_fill = PatternFill("solid", fgColor="DDEBF7")
        section_fills = {
            "Run": PatternFill("solid", fgColor="E2F0D9"),
            "Capital": PatternFill("solid", fgColor="D9EAF7"),
            "Performance": PatternFill("solid", fgColor="FFF2CC"),
            "Trades": PatternFill("solid", fgColor="E4DFEC"),
            "RSI": PatternFill("solid", fgColor="FCE4D6"),
            "Scale": PatternFill("solid", fgColor="DDEBF7"),
        }
        money_columns = {
            "entry_price", "close_price", "tactical_balance", "balance_before",
            "balance_after", "total_assets", "amount", "profit", "fee_paid",
            "save_money",
        }
        percent_columns = {
            "profit_percent", "pnl_percent", "trade_amount_percent",
            "profit_percent_per_month",
        }

        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.sheet_view.showGridLines = False
            ws.row_dimensions[1].height = 24
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            headers = {cell.value: cell.column for cell in ws[1]}
            for column_index, cells in enumerate(ws.columns, start=1):
                values = [str(cell.value or "") for cell in cells[:250]]
                width = min(55, max(11, max(map(len, values), default=10) + 2))
                ws.column_dimensions[get_column_letter(column_index)].width = width
            for name in money_columns:
                column_index = headers.get(name)
                if column_index:
                    for row_index in range(2, ws.max_row + 1):
                        ws.cell(row_index, column_index).number_format = '$#,##0.00;[Red]-$#,##0.00'
            for name in percent_columns:
                column_index = headers.get(name)
                if column_index:
                    for row_index in range(2, ws.max_row + 1):
                        ws.cell(row_index, column_index).number_format = '0.00"%";[Red]-0.00"%"'
            if ws.title == "Overview":
                section_column = headers.get("Section")
                metric_column = headers.get("Metric")
                value_column = headers.get("Value")
                for row_index in range(2, ws.max_row + 1):
                    section = str(ws.cell(row_index, section_column).value or "")
                    metric = str(ws.cell(row_index, metric_column).value or "")
                    ws.cell(row_index, section_column).font = Font(bold=True, color="1F1F1F")
                    ws.cell(row_index, section_column).fill = section_fills.get(
                        section, PatternFill("solid", fgColor="F2F2F2")
                    )
                    value_cell = ws.cell(row_index, value_column)
                    metric_lower = metric.lower()
                    if any(token in metric_lower for token in (
                        "balance", "profit", "fees", "saved money",
                    )) and "%" not in metric:
                        value_cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                    elif "%" in metric or "win rate" in metric_lower or "drawdown" in metric_lower:
                        value_cell.number_format = '0.00"%";[Red]-0.00"%"'
                    elif any(token in metric_lower for token in (
                        "trades", "wins", "losses", "positions", "liquidations", "months",
                    )):
                        value_cell.number_format = '#,##0'
            if ws.max_row >= 2:
                table = Table(
                    displayName=f"TradeReportTable{sheet_index}", ref=ws.dimensions
                )
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2", showRowStripes=True,
                    showFirstColumn=False, showLastColumn=False,
                    showColumnStripes=False,
                )
                ws.add_table(table)

            profit_column = headers.get("profit")
            close_time_column = headers.get("close_time")
            if profit_column:
                for row_index in range(2, ws.max_row + 1):
                    profit = ws.cell(row_index, profit_column).value
                    row_fill = profit_fill if isinstance(profit, (int, float)) and profit >= 0 else loss_fill
                    if close_time_column:
                        close_time = str(ws.cell(row_index, close_time_column).value)
                        if close_time in multi_close_times:
                            row_fill = grouped_fill
                    for column_index in range(1, ws.max_column + 1):
                        ws.cell(row_index, column_index).fill = row_fill

        wb.save(excel_file_name)
