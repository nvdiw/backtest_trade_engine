import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from trade_csv_logger import TradeCSVLogger


class TradeCSVLoggerWorkbookTests(unittest.TestCase):
    def test_default_workbook_has_frozen_separate_strategy_sheets(self):
        logger = TradeCSVLogger()
        logger.rows = [
            {"trade_id": "ma_strategy_0001", "type": "LONG", "profit": 10},
            {"trade_id": "rsi_ma_strategy_0002", "type": "SHORT", "profit": -2},
            {"trade_id": "scale_ma_strategy_0003", "type": "LONG", "profit": 3},
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "data_orders.csv"
            logger.save_csv(
                first_balance=1000,
                final_balance=1011,
                total_profit=11,
                total_profit_percent=1.1,
                total_fee=0.5,
                start_time="2026-01-01 00:00:00",
                end_time="2026-01-02 00:00:00",
                days=1,
                hours=0,
                minutes=0,
                overview_metrics={
                    "Run": {"Start time": "2026-01-01 00:00:00"},
                    "Capital": {"Final balance": 1011, "Total profit %": 1.1},
                    "Performance": {"Maximum drawdown %": -2.5},
                    "Trades": {"Closed trades": 3},
                    "RSI": {"RSI trades": 1},
                    "Scale": {"Scale trades": 1},
                },
                file_name=str(csv_path),
            )
            workbook = load_workbook(csv_path.with_suffix(".xlsx"), read_only=False)
            sheet_names = workbook.sheetnames
            freeze_panes = [worksheet.freeze_panes for worksheet in workbook.worksheets]
            row_counts = {
                name: workbook[name].max_row
                for name in ("Main Strategy", "RSI Strategy", "Scale Strategy")
            }
            overview_headers = [cell.value for cell in workbook["Overview"][1]]
            overview_metrics = [
                workbook["Overview"].cell(row=row, column=2).value
                for row in range(2, workbook["Overview"].max_row + 1)
            ]
            workbook.close()

        self.assertEqual(
            sheet_names,
            ["Overview", "All Trades", "Main Strategy", "RSI Strategy", "Scale Strategy"],
        )
        self.assertEqual(freeze_panes, ["A2"] * 5)
        self.assertEqual(row_counts["Main Strategy"], 2)
        self.assertEqual(row_counts["RSI Strategy"], 2)
        self.assertEqual(row_counts["Scale Strategy"], 2)
        self.assertEqual(overview_headers, ["Section", "Metric", "Value"])
        self.assertIn("Maximum drawdown %", overview_metrics)


if __name__ == "__main__":
    unittest.main()
