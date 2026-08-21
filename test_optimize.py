import csv
import gzip
import json
import re
import tempfile
import unittest
from argparse import Namespace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from optimize import (
    ExtraTreesSurrogate, SmartCandidateGenerator,
    _aggregate_walk_forward_records, _auto_bootstrap, _compact_auto_candidate_plans,
    _annotate_discovery_learning_scores, _apply_funnel_learning_scores,
    _learn_mutation_guidance, _learn_parameter_importance, _open_csv_text,
    _resolve_csv_path,
    _read_candidate_plan, _read_surrogate_history_cache,
    _representative_surrogate_history, _select_surrogate_candidates,
    _write_surrogate_history_cache,
    _robust_validation_score, _time_normalized_score,
    build_parser, grid_size,
    iter_grid_candidates,
    param_grid, run_auto_optimization, run_optimization,
)
from ma_strategy import resolve_parameter_source
from strategy_config import build_ma_strategy_config, load_ma_strategy_tune
from trade_engine import TradeEngine


class OptimizerSearchTests(unittest.TestCase):
    def test_legacy_surrogate_cache_migrates_raw_scores_to_comparable_ranks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_path = Path(temp_dir) / "surrogate_history_cache.json.gz"
            payload = {
                "version": 1,
                "latest_cycle": 12,
                "parameter_keys": ["x"],
                "rows": [["low", 100, 1], ["high", 2000, 2]],
            }
            with gzip.open(cache_path, "wt", encoding="utf-8") as cache_file:
                json.dump(payload, cache_file)

            loaded, latest_cycle = _read_surrogate_history_cache(cache_path, ("x",))

        targets = {row["candidate_id"]: row["learning_score"] for row in loaded}
        self.assertEqual(latest_cycle, 12)
        self.assertEqual(targets, {"low": 0.0, "high": 1.0})

    def test_funnel_learning_replaces_raw_scale_with_later_stage_evidence(self):
        discovery = [
            {
                "candidate_id": "lucky", "params": {"x": 1},
                "objective_score": 1000, "time_normalized_score": 1000,
            },
            {
                "candidate_id": "stable", "params": {"x": 2},
                "objective_score": 900, "time_normalized_score": 900,
            },
        ]
        _annotate_discovery_learning_scores(discovery)
        stages = {
            "discovery": discovery,
            "validation": [
                {"candidate_id": "lucky", "time_normalized_score": -50},
                {"candidate_id": "stable", "time_normalized_score": 80},
            ],
            "stress": [
                {"candidate_id": "stable", "time_normalized_score": 70},
            ],
            "walk_forward": [
                {"candidate_id": "stable", "time_normalized_score": 60},
            ],
            "final": [
                {"candidate_id": "stable", "time_normalized_score": 65},
            ],
        }

        _apply_funnel_learning_scores(discovery, stages)
        targets = {row["candidate_id"]: row["learning_score"] for row in discovery}

        self.assertGreater(targets["stable"], targets["lucky"])
        self.assertTrue(all(row["learning_source"] == "robust_funnel_rank" for row in discovery))

    def test_mutation_guidance_learns_direction_and_step_size(self):
        history = [
            {
                "candidate_id": str(value), "params": {"x": value},
                "learning_score": value / 20,
            }
            for value in range(21)
        ]

        guidance = _learn_mutation_guidance(history, ("x",), {"x": [0, 10, 20]})
        generator = SmartCandidateGenerator(
            {"x": [0.0, 10.0, 20.0]}, seed=2,
            continuous_refinement=True, mutation_guidance=guidance,
        )

        self.assertEqual(guidance["x"]["direction"], 1)
        self.assertEqual(guidance["x"]["step_multiplier"], 3)
        self.assertEqual(generator._refined_neighbors("x", 10.0)[0], 20.0)

    def test_surrogate_selection_reports_quality_uncertainty_and_diversity(self):
        grid = {"x": list(range(16)), "y": list(range(16))}
        history = [
            {
                "candidate_id": f"h-{x}-{y}", "params": {"x": x, "y": y},
                "learning_score": (x + y) / 30,
            }
            for x in range(8) for y in range(8)
        ]
        candidates = [
            {"x": x, "y": y} for x in range(16) for y in range(16)
            if x >= 8 or y >= 8
        ]
        features = {"surrogate_min_samples": 4, "surrogate_trees": 8}

        selected, metadata = _select_surrogate_candidates(
            candidates, history, 64, ("x", "y"), grid, features, seed=7,
            parameter_importance={"x": {"weight": 0.6}, "y": {"weight": 0.4}},
        )

        self.assertEqual(len(selected), 64)
        self.assertEqual(metadata["training_target"], "normalized_and_robust_funnel_rank")
        self.assertEqual(metadata["selection_mix"]["diversity_novelty"], 0.20)
        self.assertGreater(metadata["diversity_buckets"], 10)

    def test_compact_surrogate_cache_round_trips_legacy_history(self):
        history = [
            {
                "candidate_id": f"c1-{score}",
                "objective_score": score,
                "params": {"x": score, "y": score % 3},
            }
            for score in range(10_000)
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_path = Path(temp_dir) / "surrogate_history_cache.json.gz"
            selected = _write_surrogate_history_cache(
                cache_path, history, ("x", "y"), 159
            )
            loaded, latest_cycle = _read_surrogate_history_cache(
                cache_path, ("x", "y")
            )

        self.assertEqual(len(selected), 1024)
        self.assertEqual(loaded, selected)
        self.assertEqual(latest_cycle, 159)

    def test_large_surrogate_history_is_representative_and_bounded(self):
        history = [
            {"objective_score": score, "params": {"x": score}}
            for score in range(10_000)
        ]

        selected = _representative_surrogate_history(history, 1024)

        self.assertEqual(len(selected), 1024)
        self.assertEqual(selected[0]["objective_score"], 0)
        self.assertEqual(selected[-1]["objective_score"], 9_999)
        self.assertEqual(selected, _representative_surrogate_history(history, 1024))

    def test_legacy_candidate_plans_are_compacted_without_losing_candidates(self):
        candidates = [
            {"candidate_id": f"c1-{index}", "params": {"x": index, "y": index % 2}}
            for index in range(100)
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            cycle_dir = Path(temp_dir) / "cycles" / "cycle_000001"
            cycle_dir.mkdir(parents=True)
            payload = {
                "cycle": 1,
                "stage": "discovery_pool",
                "candidate_count": len(candidates),
                "candidates": candidates,
            }
            source = cycle_dir / "discovery_pool_candidates.json"
            rung = cycle_dir / "discovery_rung_01_candidates.json"
            source.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            payload["stage"] = "discovery_rung_01"
            rung.write_text(json.dumps(payload, indent=2), encoding="utf-8")

            rewritten, reclaimed = _compact_auto_candidate_plans(temp_dir)
            rung_payload = json.loads(rung.read_text(encoding="utf-8"))

            self.assertEqual(rewritten, 2)
            self.assertGreater(reclaimed, 0)
            self.assertEqual(_read_candidate_plan(source), candidates)
            self.assertEqual(_read_candidate_plan(rung), candidates)
            self.assertEqual(rung_payload["source"], "discovery_pool_candidates.json")

    def test_help_explains_modes_and_includes_runnable_examples(self):
        help_text = build_parser().format_help()

        self.assertIn("Choose one search path:", help_text)
        self.assertIn("search mode and parameter scope:", help_text)
        self.assertIn("standard search ranges and robustness:", help_text)
        self.assertIn("auto campaign (used only with --auto):", help_text)
        self.assertIn("recommended examples:", help_text)
        self.assertIn("--validation-start 2025-01-01", help_text)
        self.assertIn("--auto --auto-cycles 2", help_text)
        self.assertIn("--resume", help_text)

    def test_planning_and_output_options_are_available(self):
        args = build_parser().parse_args([
            "--profile", "risk", "--dry-run", "--top-n", "7",
            "--output-dir", "custom-output",
        ])
        self.assertEqual(args.profile, "risk")
        self.assertTrue(args.dry_run)
        self.assertEqual(args.top_n, 7)
        self.assertEqual(args.output_dir, "custom-output")

        auto_args = build_parser().parse_args(["--auto"])
        self.assertTrue(auto_args.auto)
        self.assertEqual(auto_args.auto_tests, 2000)
        self.assertEqual(auto_args.auto_validation_top, 30)
        self.assertEqual(auto_args.auto_stress_top, 10)
        self.assertEqual(auto_args.auto_final_top, 3)

    def test_grid_is_complete_and_deterministic(self):
        grid = {"a": [1, 2], "b": ["x", "y", "z"]}
        candidates = list(iter_grid_candidates(grid))

        self.assertEqual(grid_size(grid), 6)
        self.assertEqual(len(candidates), 6)
        self.assertEqual(candidates[0], {"a": 1, "b": "x"})
        self.assertEqual(candidates[-1], {"a": 2, "b": "z"})

    def test_search_grid_keeps_every_current_default_as_a_candidate(self):
        defaults = build_ma_strategy_config()
        for key, values in param_grid.items():
            self.assertIn(getattr(defaults, key), values, key)

    def test_every_config_value_used_by_ma_strategy_is_in_grid(self):
        strategy_source = Path("ma_strategy.py").read_text(encoding="utf-8-sig")
        used_config_keys = set(re.findall(r"cfg\.([A-Za-z_]\w*)", strategy_source))
        self.assertEqual(used_config_keys - set(param_grid), set())

    def test_smart_search_is_unique_and_respects_period_order(self):
        grid = {
            "ema_16_period": [10, 20],
            "ma_50_period": [15, 30],
            "ma_100_period": [25, 40],
            "ma_200_period": [35, 50],
            "entry_score_threshold": [7, 8, 9],
        }
        generator = SmartCandidateGenerator(grid, seed=7)
        candidates = generator.generate(20)
        signatures = {tuple(candidate.items()) for candidate in candidates}

        self.assertEqual(len(signatures), len(candidates))
        self.assertTrue(candidates)
        for candidate in candidates:
            periods = [
                candidate["ema_16_period"], candidate["ma_50_period"],
                candidate["ma_100_period"], candidate["ma_200_period"],
            ]
            self.assertEqual(periods, sorted(periods))

    def test_smart_search_queues_exact_neighbors_around_elites(self):
        generator = SmartCandidateGenerator(
            {"entry_score_threshold": [6, 7, 8, 9, 10]}, seed=11
        )
        generator.seen.add((8,))
        generator._queue_elite_neighbors([
            {"params": {"entry_score_threshold": 8}}
        ])

        self.assertEqual(
            [candidate["entry_score_threshold"] for candidate in generator.local_queue],
            [7, 9],
        )

    def test_auto_refinement_creates_values_between_grid_points(self):
        generator = SmartCandidateGenerator(
            {"ma_50_period": [10, 20, 30]},
            seed=11,
            continuous_refinement=True,
        )
        generator.seen.add((20,))
        generator._queue_elite_neighbors([
            {"params": {"ma_50_period": 20}}
        ])

        self.assertEqual(
            [candidate["ma_50_period"] for candidate in generator.local_queue],
            [19, 21],
        )

    def test_parameter_importance_favors_variables_with_large_score_effect(self):
        records = []
        for index in range(40):
            important = index % 2
            noise = (index // 2) % 2
            records.append({
                "params": {"important": important, "noise": noise},
                "objective_score": important * 100 + noise,
                "result": {},
            })

        learned = _learn_parameter_importance(
            records, ("important", "noise"), target="objective_score"
        )

        self.assertGreater(learned["important"]["weight"], learned["noise"]["weight"])

    def test_extra_trees_surrogate_learns_interactions_and_uncertainty(self):
        features = []
        targets = []
        for left in range(8):
            for right in range(8):
                features.append([left, right])
                targets.append(100 if left >= 5 and right <= 2 else left - right)
        model = ExtraTreesSurrogate(n_trees=24, min_leaf=2, seed=7).fit(
            features, targets
        )

        predictions = model.predict_mean_std([[6, 1], [1, 6]])

        self.assertGreater(predictions[0][0], predictions[1][0] + 40)
        self.assertGreaterEqual(predictions[0][1], 0)

    def test_walk_forward_score_penalizes_an_unstable_candidate(self):
        candidates = [
            {"candidate_id": "stable", "params": {"x": 1}},
            {"candidate_id": "unstable", "params": {"x": 2}},
        ]
        fold_records = {}
        for fold, stable_score, unstable_score in (
            ("f1", 70, 140), ("f2", 72, 140), ("f3", 71, -100),
        ):
            fold_records[fold] = [
                {
                    "candidate_id": "stable", "params": {"x": 1},
                    "objective_score": stable_score,
                    "result": {"score": stable_score, "closed_trades": 10},
                },
                {
                    "candidate_id": "unstable", "params": {"x": 2},
                    "objective_score": unstable_score,
                    "result": {"score": unstable_score, "closed_trades": 10},
                },
            ]

        aggregated = _aggregate_walk_forward_records(
            fold_records, candidates, stability_penalty=0.15
        )
        scores = {record["candidate_id"]: record["objective_score"] for record in aggregated}

        self.assertGreater(scores["stable"], scores["unstable"])

    def test_time_normalization_compares_scores_relative_to_range_length(self):
        one_month = 30 * 24 * 4
        one_year = round(365.25 * 24 * 4)

        monthly_rate = _time_normalized_score(100, one_month)
        yearly_rate = _time_normalized_score(2000, one_year)

        self.assertAlmostEqual(monthly_rate, 1217.5, places=1)
        self.assertAlmostEqual(yearly_rate, 2000, places=1)
        self.assertGreater(yearly_rate, monthly_rate)

    def test_new_auto_campaign_can_warm_start_from_existing_best_params(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            best_path = Path(temp_dir) / "best_params.json"
            best_path.write_text(
                json.dumps({"entry_score_threshold": 11}), encoding="utf-8"
            )
            args = build_parser().parse_args([
                "--auto", "--base-params", str(best_path)
            ])

            bootstrap = _auto_bootstrap(
                args, {"entry_score_threshold": [6, 7, 8, 9, 10, 11, 12]}
            )

        self.assertEqual(bootstrap["params"], {"entry_score_threshold": 11})
        self.assertEqual(bootstrap["compatible_parameter_count"], 1)

    def test_advanced_auto_reuses_history_for_surrogate_halving_and_walk_forward(self):
        args = build_parser().parse_args([
            "--auto", "--auto-tests", "64", "--auto-validation-top", "8",
            "--auto-stress-top", "4", "--auto-walk-forward-top", "3",
            "--auto-final-top", "2", "--auto-cycles", "2",
            "--auto-stress-start", "0", "--auto-validation-start", "60",
            "--auto-discovery-start", "90", "--auto-end", "120",
            "--auto-surrogate-min-samples", "4", "--workers", "1",
            "--log-every", "0", "--excel-top", "0",
        ])

        def fake_strategy(tune, start, end):
            score = 100 - abs(tune["x"] - 10) * 4 - abs(tune["y"] - 3)
            return {
                "score": score, "total_profit": score, "closed_trades": 10,
                "maximum_drawdown": -2,
            }

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=fake_strategy
            ):
                run_auto_optimization(
                    args,
                    grid={"x": list(range(16)), "y": list(range(16))},
                )
            output = Path(temp_dir)
            second_surrogate = json.loads((
                output / "cycles" / "cycle_000002" / "surrogate_search.json"
            ).read_text(encoding="utf-8"))
            state = json.loads((output / "auto_state.json").read_text(encoding="utf-8"))
            compressed_results = (
                output / "cycles" / "cycle_000001"
                / "discovery_rung_01_results.csv.gz"
            )
            with _open_csv_text(compressed_results) as results_file:
                auto_columns = next(csv.reader(results_file))

            self.assertTrue(compressed_results.is_file())
            self.assertLess(auto_columns.index("total_profit"), auto_columns.index("x"))
            self.assertTrue((
                output / "cycles" / "cycle_000001" / "walk_forward_summary.json"
            ).is_file())
            self.assertTrue(second_surrogate["enabled"])
            self.assertGreaterEqual(second_surrogate["history_samples"], 8)
            self.assertEqual(state["version"], 2)
            self.assertEqual(state["cycles_completed"], 2)

    def test_auto_mode_runs_full_funnel_and_writes_resumable_state(self):
        args = build_parser().parse_args([
            "--auto", "--auto-tests", "4", "--auto-validation-top", "3",
            "--auto-stress-top", "2", "--auto-final-top", "1",
            "--auto-cycles", "1", "--auto-stress-start", "0",
            "--auto-validation-start", "20", "--auto-discovery-start", "30",
            "--auto-end", "40", "--workers", "1", "--log-every", "0",
            "--excel-top", "1",
        ])

        def fake_strategy(tune, start, end):
            value = tune["x"]
            return {
                "score": value,
                "total_profit": value * 10,
                "total_profit_percent": value,
                "closed_trades": 10,
                "maximum_drawdown": -1,
            }

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=fake_strategy
            ) as strategy:
                best = run_auto_optimization(args, grid={"x": [1, 2, 3, 4]})
            state = json.loads(
                (Path(temp_dir) / "auto_state.json").read_text(encoding="utf-8")
            )
            saved = json.loads(
                (Path(temp_dir) / "best_params.json").read_text(encoding="utf-8")
            )
            workbook = load_workbook(
                Path(temp_dir) / "auto_report.xlsx", read_only=True
            )
            auto_sheets = workbook.sheetnames
            workbook.close()
            completed_checkpoints_removed = not (
                Path(temp_dir) / "cycles" / "cycle_000001" / "checkpoints"
            ).exists()

        self.assertEqual(strategy.call_count, 10)
        self.assertEqual(state["cycles_completed"], 1)
        self.assertEqual(state["status"], "completed")
        self.assertEqual(state["config"]["profile"], "full")
        self.assertEqual(saved, {"x": 4})
        self.assertEqual(best["params"], {"x": 4})
        self.assertEqual(auto_sheets, ["Hall of Fame", "Parameter Importance"])
        self.assertTrue(completed_checkpoints_removed)

    def test_next_auto_cycle_continues_from_hall_of_fame_winner(self):
        args = build_parser().parse_args([
            "--auto", "--auto-tests", "4", "--auto-validation-top", "1",
            "--auto-stress-top", "1", "--auto-final-top", "1",
            "--auto-cycles", "2", "--auto-stress-start", "0",
            "--auto-validation-start", "20", "--auto-discovery-start", "30",
            "--auto-end", "40", "--workers", "1", "--log-every", "0",
            "--excel-top", "0",
        ])

        def fake_strategy(tune, start, end):
            value = tune["x"]
            score = 100 - abs(value - 20)
            return {
                "score": score,
                "total_profit": score,
                "closed_trades": 2,
                "maximum_drawdown": -1,
            }

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=fake_strategy
            ):
                run_auto_optimization(args, grid={"x": [0, 10, 20, 30]})
            parent = json.loads((
                Path(temp_dir) / "cycles" / "cycle_000002" / "training_parent.json"
            ).read_text(encoding="utf-8"))
            second_plan = _read_candidate_plan(
                Path(temp_dir) / "cycles" / "cycle_000002"
                / "discovery_candidates.json"
            )

        refined_values = {
            candidate["params"]["x"] for candidate in second_plan
        }
        self.assertEqual(parent["source"], "hall_of_fame")
        self.assertEqual(parent["params"]["x"], 20)
        self.assertTrue(refined_values & {19, 21})

    def test_auto_resume_continues_the_interrupted_stage(self):
        args = build_parser().parse_args([
            "--auto", "--auto-tests", "2", "--auto-validation-top", "2",
            "--auto-stress-top", "1", "--auto-final-top", "1",
            "--auto-cycles", "1", "--auto-stress-start", "0",
            "--auto-validation-start", "20", "--auto-discovery-start", "30",
            "--auto-end", "40", "--workers", "1", "--log-every", "0",
            "--excel-top", "0",
        ])

        def fake_strategy(tune, start, end):
            value = tune["x"]
            return {
                "score": value,
                "total_profit": value,
                "closed_trades": 2,
                "maximum_drawdown": -1,
            }

        interrupted_calls = 0

        def interrupt_after_one_result(tune, start, end):
            nonlocal interrupted_calls
            interrupted_calls += 1
            if interrupted_calls == 2:
                raise KeyboardInterrupt
            return fake_strategy(tune, start, end)

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=interrupt_after_one_result
            ):
                first = run_auto_optimization(args, grid={"x": [1, 2]})
            interrupted_state = json.loads(
                (Path(temp_dir) / "auto_state.json").read_text(encoding="utf-8")
            )
            interrupted_best_exists = (
                Path(temp_dir) / "best_params.json"
            ).is_file() and (
                Path(temp_dir) / "cycles" / "cycle_000001" / "checkpoints"
                / "discovery" / "best_params.json"
            ).is_file()

            # Simulate an interrupted version-1 campaign. Plain --auto must
            # migrate and resume it without requiring an explicit --resume.
            state_path = Path(temp_dir) / "auto_state.json"
            legacy_state = json.loads(state_path.read_text(encoding="utf-8"))
            legacy_state["version"] = 1
            legacy_state.pop("optimizer_features", None)
            legacy_state.pop("advanced_from_cycle", None)
            state_path.write_text(json.dumps(legacy_state), encoding="utf-8")
            discovery_path = (
                Path(temp_dir) / "cycles" / "cycle_000001"
                / "discovery_results.csv"
            )
            with discovery_path.open(newline="", encoding="utf-8") as csv_file:
                legacy_rows = list(csv.DictReader(csv_file))
            legacy_fields = [
                key for key in legacy_rows[0]
                if key not in ("time_normalized_score", "range_candles")
            ]
            with discovery_path.open("w", newline="", encoding="utf-8") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=legacy_fields)
                writer.writeheader()
                writer.writerows({key: row[key] for key in legacy_fields} for row in legacy_rows)
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=fake_strategy
            ):
                resumed = run_auto_optimization(args, grid={"x": [1, 2]})
            completed_state = json.loads(
                (Path(temp_dir) / "auto_state.json").read_text(encoding="utf-8")
            )
            discovery_results = _resolve_csv_path(
                Path(temp_dir) / "cycles" / "cycle_000001" / "discovery_results.csv"
            )
            with _open_csv_text(discovery_results) as results_file:
                discovery_reader = csv.DictReader(results_file)
                discovery_rows = list(discovery_reader)
                discovery_fields = discovery_reader.fieldnames

        self.assertIsNone(first)
        self.assertEqual(interrupted_state["status"], "interrupted")
        self.assertEqual(interrupted_state["stage_completed"], 1)
        self.assertTrue(interrupted_best_exists)
        self.assertEqual(completed_state["status"], "completed")
        self.assertEqual(completed_state["version"], 2)
        self.assertEqual(completed_state["migrated_from_version"], 1)
        self.assertEqual(completed_state["cycles_completed"], 1)
        self.assertEqual(completed_state["total_evaluations"], 6)
        self.assertEqual(discovery_results.suffixes[-2:], [".csv", ".gz"])
        self.assertEqual(len(discovery_rows), 2)
        self.assertIn("time_normalized_score", discovery_fields)
        self.assertIn("range_candles", discovery_fields)
        self.assertLess(discovery_fields.index("total_profit"), discovery_fields.index("x"))
        self.assertEqual(len({row["candidate_id"] for row in discovery_rows}), 2)
        self.assertEqual(resumed["params"], {"x": 2})

    def test_inactive_filter_parameters_collapse_to_one_effective_signature(self):
        generator = SmartCandidateGenerator({
            "volume_filter": [False, True],
            "volume_spike_multiplier": [1.0, 1.5],
            "entry_score_volume": [1, 2],
        })
        first = generator._canonicalize({
            "volume_filter": False,
            "volume_spike_multiplier": 1.0,
            "entry_score_volume": 1,
        })
        second = generator._canonicalize({
            "volume_filter": False,
            "volume_spike_multiplier": 1.5,
            "entry_score_volume": 2,
        })

        self.assertEqual(generator._signature(first), generator._signature(second))

    def test_grid_run_writes_best_params_as_json_by_score(self):
        args = Namespace(
            output_dir=None, mode="grid", tests=99, workers=1, batch_size=2,
            chunksize=1, start="0", end="10", seed=3, log_every=0,
            elite_size=2,
        )

        def fake_strategy(tune, start, end):
            value = tune["x"]
            return {"score": value, "total_profit": 100 - value}

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch("optimize.ma_strategy", fake_strategy):
                best = run_optimization(args, grid={"x": [1, 2, 3]})
            with (Path(temp_dir) / "best_params.json").open(encoding="utf-8") as file:
                saved = json.load(file)
            workbook = load_workbook(
                Path(temp_dir) / "optimization_results.xlsx", read_only=True
            )
            with (Path(temp_dir) / "optimization_results.csv").open(
                encoding="utf-8"
            ) as results_file:
                columns = next(csv.reader(results_file))
            sheet_names = workbook.sheetnames
            workbook.close()

        self.assertEqual(best["params"], {"x": 3})
        self.assertEqual(saved, {"x": 3})
        self.assertIn("Core Metrics", sheet_names)
        self.assertIn("RSI Metrics", sheet_names)
        self.assertIn("Scale Metrics", sheet_names)
        self.assertIn("objective_score", columns)
        self.assertIn("profit_per_trade", columns)
        self.assertLess(columns.index("total_profit"), columns.index("x"))

    def test_keyboard_interrupt_stops_cleanly_and_keeps_checkpoint(self):
        args = Namespace(
            output_dir=None, mode="smart", tests=3, workers=1, batch_size=2,
            chunksize=1, start="0", end="10", seed=3, log_every=0,
            elite_size=2,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=KeyboardInterrupt
            ):
                best = run_optimization(args, grid={"entry_score_threshold": [7, 8]})
            results_path = Path(temp_dir) / "optimization_results.csv"
            rows = results_path.read_text(encoding="utf-8").splitlines()

        self.assertIsNone(best)
        self.assertEqual(len(rows), 1)


class PerformanceScoreTests(unittest.TestCase):
    def score(self, **overrides):
        values = {
            "return_percent": 20,
            "max_drawdown": -10,
            "win_rate": 60,
            "profits": [10, -5] * 20,
            "first_balance": 1000,
            "profit_months": 8,
            "loss_months": 4,
            "liquidations": 0,
        }
        values.update(overrides)
        return TradeEngine.calculate_performance_score(**values)

    def test_no_trade_result_cannot_win(self):
        metrics = self.score(profits=[], win_rate=0)
        self.assertEqual(metrics["score"], -1_000_000.0)

    def test_drawdown_and_liquidation_reduce_score(self):
        safe = self.score(max_drawdown=-5)
        risky = self.score(max_drawdown=-30, liquidations=8)
        self.assertGreater(safe["score"], risky["score"])

    def test_score_exposes_quality_metrics(self):
        metrics = self.score()
        self.assertEqual(metrics["profit_factor"], 2.0)
        self.assertIn("expectancy_percent", metrics)
        self.assertIn("calmar_ratio", metrics)


class ParameterSourceTests(unittest.TestCase):
    def test_config_and_best_sources_are_explicit(self):
        tune, description = resolve_parameter_source("config")
        self.assertIsNone(tune)
        self.assertEqual(description, "strategy_config.py")

        with tempfile.TemporaryDirectory() as temp_dir:
            best_path = Path(temp_dir) / "best.json"
            best_path.write_text('{"leverage": 7, "fee_rate": 0.001}', encoding="utf-8")
            tune, description = resolve_parameter_source("best", best_params=best_path)

        self.assertEqual(tune["leverage"], 7)
        self.assertEqual(tune["fee_rate"], 0.001)
        self.assertEqual(description, str(best_path))

    def test_parameter_json_rejects_unknown_names(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "bad.json"
            path.write_text('{"leverge": 7}', encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "leverge"):
                load_ma_strategy_tune(path)

    def test_optimizer_preserves_base_parameters_outside_profile(self):
        args = Namespace(
            output_dir=None, mode="grid", tests=1, workers=1, batch_size=2,
            chunksize=1, start="0", end="10", seed=3, log_every=0,
            elite_size=2, base_source="file", base_params=None,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir
            args.base_params = str(Path(temp_dir) / "base.json")
            Path(args.base_params).write_text('{"leverage": 7}', encoding="utf-8")

            def fake_strategy(tune, start, end):
                self.assertEqual(tune["leverage"], 7)
                return {"score": tune["ema_16_period"]}

            with patch("optimize._init_worker"), patch("optimize.ma_strategy", fake_strategy):
                best = run_optimization(args, grid={"ema_16_period": [10]})

        self.assertEqual(best["params"]["leverage"], 7)
        self.assertEqual(best["params"]["ema_16_period"], 10)

    def test_resume_does_not_repeat_completed_grid_candidates(self):
        args = Namespace(
            output_dir=None, mode="grid", tests=2, workers=1, batch_size=2,
            chunksize=1, start="0", end="10", seed=3, log_every=0,
            elite_size=2, resume=False,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            args.output_dir = temp_dir

            def fake_strategy(tune, start, end):
                return {"score": tune["ema_16_period"]}

            with patch("optimize._init_worker"), patch("optimize.ma_strategy", fake_strategy):
                run_optimization(args, grid={"ema_16_period": [10, 12]})

            args.resume = True
            with patch("optimize._init_worker"), patch(
                "optimize.ma_strategy", side_effect=AssertionError("candidate repeated")
            ):
                run_optimization(args, grid={"ema_16_period": [10, 12]})

            with (Path(temp_dir) / "optimization_results.csv").open(
                encoding="utf-8"
            ) as results_file:
                rows = list(csv.DictReader(results_file))

        self.assertEqual(len(rows), 2)

    def test_validation_score_penalizes_train_only_performance(self):
        stable = _robust_validation_score({"score": 100}, {"score": 90}, 0.5)
        overfit = _robust_validation_score({"score": 200}, {"score": 90}, 0.5)
        self.assertGreater(stable, overfit)


if __name__ == "__main__":
    unittest.main()
